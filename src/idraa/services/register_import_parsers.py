"""Register import parsers — xlsx (openpyxl, read-only) + CSV, one output shape.

Epic #34 P1c (register import UI): a "register" is an arbitrary spreadsheet
the operator uploads — headers and content are NOT known ahead of time (unlike
``scenario_import_parsers``'s fixed ``CSV_HEADERS``). This module only does
STRUCTURAL extraction: sniff the format, list xlsx sheet names, and parse
into a flat ``ParsedRegister`` (headers + row dicts). Column mapping,
value-binding, and any FAIR-CAM interpretation live downstream in
``register_import`` (Task 3) — this module has zero knowledge of
scenario/risk semantics.

xlsx is openpyxl's FIRST runtime read surface in this codebase — P1c promotes
openpyxl from the dev extra to a genuine ``[project]`` dependency (see the
pyproject.toml comment on that line). Untrusted xlsx bytes are therefore a
real attack surface (zip bombs, XML entity expansion):

- Every xlsx entry point below runs ``_zip_guard`` BEFORE ``load_workbook`` —
  ``list_sheet_names`` AND ``parse_register`` are both independently
  reachable from the route layer, so both must guard (Sec-N plan-gate
  amendment). The guard reads only zip central-directory metadata
  (``ZipInfo.file_size`` / ``compress_size`` / member count) and never
  decompresses to check it. It caps member count, per-member declared size, and
  per-member compression ratio (the zip-bomb signature). ``zipfile`` bounds
  every read to the declared ``file_size``, so those declared sizes are a safe
  upper bound on actual extraction; worksheet sheet-data is streamed and media
  is never read, so the only memory amplifier — the fully-buffered parts
  (sharedStrings / styles) — is bounded per-part by the per-member cap. See the
  constant block below and docs/security/threat-model.md B9 for the residual.
- Entity expansion (e.g. "billion laughs") is a SEPARATE attack from a zip
  bomb — a tiny, valid zip member can still carry a malicious XML entity
  declaration. openpyxl auto-detects ``defusedxml`` (an explicit runtime
  dependency alongside openpyxl — see pyproject.toml) and substitutes it for
  stdlib ``xml.etree`` parsing, which blocks entity declarations outright.
  ``defusedxml`` is added explicitly (rather than left to arrive
  incidentally as a transitive dev-only dependency of jupyter/cyclonedx-bom,
  which is ABSENT from a ``--no-dev`` production install) so this guarantee
  holds in production, not just in the dev/test venv.
"""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass
from typing import Any

import openpyxl

from idraa.services.scenario_import_parsers import MAX_ROWS

__all__ = [
    "MAX_ROWS",
    "ParsedRegister",
    "list_sheet_names",
    "parse_register",
    "sniff_register_format",
]

_XLSX_MAGIC = b"PK\x03\x04"

# Reject a zip/xlsx payload shaped like a zip bomb before ``load_workbook`` —
# checked purely from zip central-directory metadata, never decompressing.
# ``zipfile`` bounds every ``read()`` to the member's declared ``file_size``
# (verified: a 210 MB member yields exactly 210 MB, never more), so declared
# sizes ARE the extraction bound. What that bound protects:
#   * openpyxl (read_only) STREAMS worksheet sheet-data (a 120k-row sheet costs
#     ~1 MB RSS) and never reads media, so a large SHEET is cheap — an aggregate
#     size cap would only false-reject legit sheet-heavy workbooks.
#   * The one memory amplifier is the fully-buffered parts openpyxl loads whole
#     (sharedStrings / styles; ~7.6x RSS), and each is already bounded to 50 MB
#     by the per-member cap (~380 MB RSS per part). A tighter, non-bypassable
#     bound would have to identify those parts by ROLE — their PATH is
#     attacker-controlled via ``[Content_Types].xml`` (openpyxl resolves
#     sharedStrings by manifest PartName, not filename), so any path-based
#     heuristic is trivially bypassed. Parsing the manifest to do that is
#     deferred as disproportionate for an ADMIN-only import surface whose
#     residual (~380 MB, up to ~760 MB if both sharedStrings and styles are
#     maxed) is survivable on the 4 GB VM. See docs/security/threat-model.md B9.
_ZIP_BOMB_MAX_MEMBER_BYTES = 50 * 1024 * 1024
_ZIP_BOMB_MAX_MEMBERS = 200
# Per-member compression-ratio ceiling — a tripwire for degenerate/accidental
# bombs, not a bound on a determined attacker. Real xlsx XML compresses ~5-23x;
# valid structured XML maxes ~412x for realistic repeated XML tokens (deflate
# reaches its ~1032x ceiling only on all-same-byte runs), so 500x is
# reached only by all-same-byte members (which openpyxl would anyway reject as
# invalid XML, or never read). It is kept because it is the one clean,
# non-bypassable, false-positive-free check available, and it fails an accidental
# all-zeros upload cleanly. It does NOT bound the real residual — 5 MB of wire
# reaches the 50 MB per-member limit at any ratio (see the note above and B9).
# Applied above an absolute floor so tiny high-ratio members (harmless in
# absolute terms) don't false-positive.
_ZIP_BOMB_MAX_RATIO = 500
_ZIP_BOMB_RATIO_FLOOR_BYTES = 1 * 1024 * 1024


@dataclass
class ParsedRegister:
    """Structural extraction of one register file — no semantic interpretation.

    ``rows`` values are ``header -> str(cell)`` (every cell coerced via
    ``str(v).strip()``, ``None`` -> ``""``) plus a synthetic ``"_row"`` key
    holding the 1-based source row number (the physical spreadsheet/CSV row,
    counting the header as row 1) for error messages and dedup keys
    downstream.
    """

    headers: list[str]
    rows: list[dict[str, str]]


def _zip_guard(data: bytes) -> None:
    """Reject an xlsx payload shaped like a zip bomb, before ``load_workbook``.

    Runs before EVERY xlsx entry point in this module (Sec-N: both
    ``list_sheet_names`` and ``parse_register`` are independently reachable
    from the route layer). Reads only ``zipfile.ZipInfo`` central-directory
    metadata — never decompresses a member to check it.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            infos = zf.infolist()
    except zipfile.BadZipFile as exc:
        raise ValueError(f"workbook rejected: not a valid xlsx (zip) container: {exc}") from exc

    if len(infos) > _ZIP_BOMB_MAX_MEMBERS:
        raise ValueError(
            f"workbook rejected: {len(infos)} zip members exceeds the "
            f"{_ZIP_BOMB_MAX_MEMBERS} member cap"
        )
    oversize = [i.filename for i in infos if i.file_size > _ZIP_BOMB_MAX_MEMBER_BYTES]
    if oversize:
        raise ValueError(
            f"workbook rejected: zip member(s) declare more than "
            f"{_ZIP_BOMB_MAX_MEMBER_BYTES} uncompressed bytes: {oversize}"
        )
    for i in infos:
        if (
            i.compress_size > 0
            and i.file_size > _ZIP_BOMB_RATIO_FLOOR_BYTES
            and i.file_size / i.compress_size > _ZIP_BOMB_MAX_RATIO
        ):
            raise ValueError(
                f"workbook rejected: zip member {i.filename!r} compression "
                f"ratio {i.file_size / i.compress_size:.0f}x exceeds the "
                f"{_ZIP_BOMB_MAX_RATIO}x cap"
            )


def sniff_register_format(filename: str, content_type: str | None, data: bytes) -> str:
    """Return ``"xlsx"`` or ``"csv"``; raise ``ValueError`` on a strong conflict.

    Order: (1) filename extension, (2) content-type, (3) content peek (the
    xlsx zip magic ``PK\\x03\\x04``) — mirrors
    ``scenario_import_parsers.sniff_format``'s shape. ``filename`` is a real
    string here: ``UploadFile.filename`` is ``str | None`` at the route, but
    None/empty is rejected with a 422 BEFORE this function is ever called
    (Sec-N plan-gate amendment) — this parser API only deals in real strings.
    """
    looks_xlsx = data[:4] == _XLSX_MAGIC

    ext: str | None = None
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        ext = "xlsx"
    elif lower.endswith(".csv"):
        ext = "csv"

    if ext == "csv" and looks_xlsx:
        raise ValueError("format conflict: .csv file but content looks like an xlsx (zip) archive")
    if ext == "xlsx" and not looks_xlsx:
        raise ValueError("format conflict: .xlsx file but content is not a valid zip/xlsx archive")
    if ext is not None:
        return ext

    if content_type:
        ct = content_type.lower()
        if "spreadsheet" in ct or "excel" in ct:
            return "xlsx"
        if "csv" in ct:
            return "csv"

    return "xlsx" if looks_xlsx else "csv"


def list_sheet_names(data: bytes) -> list[str]:
    """Return the workbook's sheet names, in file order. xlsx only."""
    _zip_guard(data)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _coerce_cell(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _parse_xlsx(data: bytes, sheet_name: str | None) -> ParsedRegister:
    _zip_guard(data)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        try:
            ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        except KeyError as exc:
            raise ValueError(f"workbook rejected: no sheet named {sheet_name!r}") from exc

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("workbook rejected: sheet has no header row") from None

        headers = [_coerce_cell(h) for h in header_row]
        rows: list[dict[str, str]] = []
        for source_row, raw_row in enumerate(rows_iter, start=2):
            cells = [_coerce_cell(v) for v in (raw_row or ())]
            if not any(cells):
                continue  # skip fully-empty rows
            if len(rows) >= MAX_ROWS:
                raise ValueError(f"too many rows: maximum {MAX_ROWS} per upload")
            padded = cells + [""] * (len(headers) - len(cells))
            row = dict(zip(headers, padded[: len(headers)], strict=True))
            row["_row"] = str(source_row)
            rows.append(row)
        return ParsedRegister(headers=headers, rows=rows)
    finally:
        wb.close()


def _parse_csv(data: bytes) -> ParsedRegister:
    try:
        text = data.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError as exc:
        raise ValueError(f"file is not valid UTF-8: {exc}") from exc

    reader = csv.reader(io.StringIO(text))
    rows_raw = list(reader)
    indexed = [
        (i, cells) for i, cells in enumerate(rows_raw, start=1) if any(c.strip() for c in cells)
    ]
    if not indexed:
        raise ValueError("CSV is empty or has no header row")

    _, header_cells = indexed[0]
    headers = [c.strip() for c in header_cells]

    data_rows = indexed[1:]
    if len(data_rows) > MAX_ROWS:
        raise ValueError(f"too many rows: maximum {MAX_ROWS} per upload")

    rows: list[dict[str, str]] = []
    for source_row, cells in data_rows:
        padded = [c.strip() for c in cells] + [""] * (len(headers) - len(cells))
        row = dict(zip(headers, padded[: len(headers)], strict=True))
        row["_row"] = str(source_row)
        rows.append(row)
    return ParsedRegister(headers=headers, rows=rows)


def parse_register(data: bytes, fmt: str, sheet_name: str | None) -> ParsedRegister:
    """Parse register ``data`` -> ``ParsedRegister``. ``fmt`` is ``"xlsx"`` or ``"csv"``."""
    if fmt == "xlsx":
        return _parse_xlsx(data, sheet_name)
    if fmt == "csv":
        return _parse_csv(data)
    raise ValueError(f"unknown register format: {fmt!r}")
