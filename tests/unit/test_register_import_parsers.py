"""Register import parsers — structural extraction only (epic #34 P1c Task 1).

Covers: format sniffing, xlsx sheet listing, xlsx/CSV parsing into the flat
``ParsedRegister`` shape, and the xlsx hardening surface (zip-bomb guard,
row cap, entity-expansion safety) — openpyxl's FIRST runtime read in this
codebase, so untrusted bytes get real scrutiny before ``load_workbook``.
"""

from __future__ import annotations

import io
import struct
import zipfile

import openpyxl
import pytest

from idraa.services.register_import_parsers import (
    ParsedRegister,
    _zip_guard,
    list_sheet_names,
    parse_register,
    sniff_register_format,
)
from idraa.services.scenario_import_parsers import MAX_ROWS

# ---------------------------------------------------------------------------
# xlsx fixture builders (tests may use openpyxl write-mode freely)
# ---------------------------------------------------------------------------


def _xlsx_bytes(sheets: dict[str, list[list[object]]]) -> bytes:
    """Build a minimal in-memory xlsx from ``{sheet_name: [[row...], ...]}``."""
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        if first:
            ws.title = name
            first = False
        for row in rows:
            ws.append(row)
    wb.save(buf)
    return buf.getvalue()


def _forge_zip_member_size(data: bytes, fake_size: int) -> bytes:
    """Patch the FIRST member's declared uncompressed size in both the local
    file header and the central directory record, without writing ``fake_size``
    bytes of real payload — the zip-bomb guard reads this metadata via
    ``zipfile.ZipInfo.file_size`` and must never decompress to check it, so a
    forged declared size (real payload stays tiny) is exactly what it must catch.
    """
    buf = bytearray(data)
    lfh_off = buf.index(b"PK\x03\x04")
    cdh_off = buf.index(b"PK\x01\x02")
    struct.pack_into("<I", buf, lfh_off + 22, fake_size)
    struct.pack_into("<I", buf, cdh_off + 24, fake_size)
    return bytes(buf)


# ---------------------------------------------------------------------------
# sniff_register_format
# ---------------------------------------------------------------------------


def test_sniff_by_extension_xlsx() -> None:
    data = _xlsx_bytes({"Sheet1": [["a"]]})
    assert sniff_register_format(filename="x.xlsx", content_type=None, data=data) == "xlsx"


def test_sniff_by_extension_csv() -> None:
    assert sniff_register_format(filename="x.csv", content_type=None, data=b"a,b\n") == "csv"


def test_sniff_by_content_type_xlsx() -> None:
    data = _xlsx_bytes({"Sheet1": [["a"]]})
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert sniff_register_format(filename="upload", content_type=ct, data=data) == "xlsx"


def test_sniff_by_content_type_csv() -> None:
    assert sniff_register_format(filename="upload", content_type="text/csv", data=b"a,b\n") == "csv"


def test_sniff_defaults_to_csv_on_no_signal() -> None:
    assert sniff_register_format(filename="upload", content_type=None, data=b"a,b\n") == "csv"


def test_sniff_conflict_csv_extension_but_xlsx_magic_raises() -> None:
    data = _xlsx_bytes({"Sheet1": [["a"]]})
    with pytest.raises(ValueError, match="conflict"):
        sniff_register_format(filename="register.csv", content_type=None, data=data)


def test_sniff_conflict_xlsx_extension_without_magic_raises() -> None:
    with pytest.raises(ValueError, match="conflict"):
        sniff_register_format(filename="register.xlsx", content_type=None, data=b"name,desc\n")


# ---------------------------------------------------------------------------
# list_sheet_names
# ---------------------------------------------------------------------------


def test_list_sheet_names_preserves_file_order() -> None:
    data = _xlsx_bytes({"First": [["a"]], "Second": [["b"]], "Third": [["c"]]})
    assert list_sheet_names(data) == ["First", "Second", "Third"]


def test_list_sheet_names_runs_zip_guard() -> None:
    data = _xlsx_bytes({"Sheet1": [["a"]]})
    forged = _forge_zip_member_size(data, 200 * 1024 * 1024)
    with pytest.raises(ValueError, match="workbook rejected"):
        list_sheet_names(forged)


# ---------------------------------------------------------------------------
# parse_register — xlsx happy paths
# ---------------------------------------------------------------------------


def test_parse_xlsx_happy_three_rows() -> None:
    data = _xlsx_bytes(
        {
            "Sheet1": [
                ["Title", "Likelihood", "Impact"],
                ["Phishing", "High", "Severe"],
                ["Malware", "Medium", "Moderate"],
                ["Insider", "Low", "Minor"],
            ]
        }
    )
    parsed = parse_register(data, "xlsx", None)
    assert isinstance(parsed, ParsedRegister)
    assert parsed.headers == ["Title", "Likelihood", "Impact"]
    assert len(parsed.rows) == 3
    assert parsed.rows[0] == {
        "Title": "Phishing",
        "Likelihood": "High",
        "Impact": "Severe",
        "_row": "2",
    }
    assert parsed.rows[1]["_row"] == "3"
    assert parsed.rows[2]["_row"] == "4"


def test_parse_xlsx_selects_named_sheet() -> None:
    data = _xlsx_bytes(
        {
            "First": [["a"], ["1"]],
            "Second": [["Title"], ["Ransomware"]],
        }
    )
    parsed = parse_register(data, "xlsx", "Second")
    assert parsed.headers == ["Title"]
    assert parsed.rows == [{"Title": "Ransomware", "_row": "2"}]


def test_parse_xlsx_skips_fully_empty_rows() -> None:
    data = _xlsx_bytes(
        {
            "Sheet1": [
                ["Title"],
                ["Phishing"],
                [None],
                [""],
                ["Malware"],
            ]
        }
    )
    parsed = parse_register(data, "xlsx", None)
    assert [r["Title"] for r in parsed.rows] == ["Phishing", "Malware"]


def test_parse_xlsx_formula_cell_never_leaks_formula_text() -> None:
    # data_only=True reads the CACHED value; a workbook written by openpyxl
    # itself never stores one for formula cells, so the coerced cell must be
    # "" — never the literal "=1+1" formula string.
    data = _xlsx_bytes({"Sheet1": [["Title", "Calc"], ["Phishing", "=1+1"]]})
    parsed = parse_register(data, "xlsx", None)
    assert parsed.rows[0]["Calc"] == ""
    assert "=" not in parsed.rows[0]["Calc"]


def test_parse_xlsx_over_max_rows_raises() -> None:
    rows = [["Title"]] + [[f"Scenario {i}"] for i in range(MAX_ROWS + 1)]
    data = _xlsx_bytes({"Sheet1": rows})
    with pytest.raises(ValueError, match="too many rows"):
        parse_register(data, "xlsx", None)


def test_parse_xlsx_unknown_sheet_name_raises() -> None:
    data = _xlsx_bytes({"Sheet1": [["a"]]})
    with pytest.raises(ValueError, match="Nonexistent"):
        parse_register(data, "xlsx", "Nonexistent")


def test_parse_xlsx_sheet_with_no_rows_raises() -> None:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active.title = "Empty"  # never appended to — zero rows, not even a header
    wb.save(buf)
    with pytest.raises(ValueError, match="no header row"):
        parse_register(buf.getvalue(), "xlsx", None)


# ---------------------------------------------------------------------------
# parse_register — xlsx hardening (zip guard + entity expansion)
# ---------------------------------------------------------------------------


def test_parse_xlsx_runs_zip_guard_before_load() -> None:
    # Binding plan-gate amendment: parse_register (not only list_sheet_names)
    # must run _zip_guard before load_workbook.
    data = _xlsx_bytes({"Sheet1": [["a"]]})
    forged = _forge_zip_member_size(data, 200 * 1024 * 1024)
    with pytest.raises(ValueError, match="workbook rejected"):
        parse_register(forged, "xlsx", None)


def test_zip_guard_rejects_oversize_member_directly() -> None:
    data = _xlsx_bytes({"Sheet1": [["a"]]})
    forged = _forge_zip_member_size(data, 200 * 1024 * 1024)
    with pytest.raises(ValueError, match="workbook rejected"):
        _zip_guard(forged)


def test_zip_guard_rejects_too_many_members() -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for i in range(201):
            zf.writestr(f"member{i}.bin", b"x")
    with pytest.raises(ValueError, match="workbook rejected"):
        _zip_guard(buf.getvalue())


def test_zip_guard_rejects_not_a_zip() -> None:
    with pytest.raises(ValueError, match="workbook rejected"):
        _zip_guard(b"not a zip file at all")


def test_parse_xlsx_entity_declaration_never_expands(tmp_path) -> None:
    """Security-shaped (Sec-N): openpyxl+defusedxml must be active — a workbook
    part carrying an XML entity declaration must never be expanded; the parser
    fails clean and fast rather than hanging or leaking the expanded text.

    A dev venv with lxml present takes openpyxl's lxml safe-parser branch
    (``resolve_entities=False``) instead of defusedxml's — both are entity-safe
    backends, so this test pins the OBSERVABLE guarantee (entity declarations
    are never honoured) that holds regardless of which backend is active. In
    production (no lxml — dev-only transitive dep), defusedxml is what
    provides it, which is why defusedxml is now an explicit runtime dependency
    (pyproject.toml) rather than an incidental one.
    """
    data = _xlsx_bytes({"Sheet1": [["Title"], ["placeholder"]]})
    buf = bytearray(data)
    zin = zipfile.ZipFile(io.BytesIO(bytes(buf)))
    malicious_sheet = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        b'<!DOCTYPE worksheet [<!ENTITY xxe "pwned">]>\n'
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row r="1"><c r="A1" t="inlineStr">'
        b"<is><t>&xxe;</t></is></c></row></sheetData></worksheet>"
    )
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w") as zout:
        for item in zin.infolist():
            content = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                content = malicious_sheet
            zout.writestr(item, content)
    malicious = out.getvalue()

    with pytest.raises(ValueError):
        parse_register(malicious, "xlsx", None)


# ---------------------------------------------------------------------------
# parse_register — CSV
# ---------------------------------------------------------------------------


def test_parse_csv_happy_path() -> None:
    data = b"Title,Likelihood\nPhishing,High\nMalware,Medium\n"
    parsed = parse_register(data, "csv", None)
    assert parsed.headers == ["Title", "Likelihood"]
    assert parsed.rows == [
        {"Title": "Phishing", "Likelihood": "High", "_row": "2"},
        {"Title": "Malware", "Likelihood": "Medium", "_row": "3"},
    ]


def test_parse_csv_strips_utf8_bom() -> None:
    data = b"\xef\xbb\xbfTitle,Likelihood\nPhishing,High\n"
    parsed = parse_register(data, "csv", None)
    assert parsed.headers == ["Title", "Likelihood"]


def test_parse_csv_skips_fully_empty_rows() -> None:
    data = b"Title\nPhishing\n\nMalware\n"
    parsed = parse_register(data, "csv", None)
    assert [r["Title"] for r in parsed.rows] == ["Phishing", "Malware"]


def test_parse_csv_over_max_rows_raises() -> None:
    body = "Title\n" + "\n".join(f"Scenario {i}" for i in range(MAX_ROWS + 1)) + "\n"
    with pytest.raises(ValueError, match="too many rows"):
        parse_register(body.encode("utf-8"), "csv", None)


def test_parse_csv_empty_file_raises() -> None:
    with pytest.raises(ValueError, match="empty"):
        parse_register(b"", "csv", None)


def test_parse_csv_invalid_utf8_raises() -> None:
    with pytest.raises(ValueError, match="not valid UTF-8"):
        parse_register(b"Title\n\xff\xfe not utf-8\n", "csv", None)


def test_parse_register_unknown_format_raises() -> None:
    with pytest.raises(ValueError, match="unknown register format"):
        parse_register(b"a,b\n", "yaml", None)
