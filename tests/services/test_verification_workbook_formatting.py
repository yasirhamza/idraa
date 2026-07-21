"""Layout + formatting tests for the verification workbook (value-preserving).

This characterization test LOCKS the FAIR math (App numeric values + LET formula
strings) before any styling/reorder pass. Two invariants are coordinate-free
(survive a column/row reorder):

  (a) the multiset of App numeric values on the MC sheet, and
  (b) the self-contained LET formula strings (they reference RANDARRAY arrays,
      not sheet cells, so they carry no coordinates).

NOTE ON RECONSTRUCTIBILITY: ``_make_completed_single_run`` uses
``_controls_snapshot_v2`` controls (snapshot_version=2), which raises
``LegacySnapshotError`` inside ``snapshot_to_fair_cam_controls``. This prevents
the LET from being emitted on the residual path. To get a real LET we call
``_make_completed_single_run`` with ``controls=[]`` — an empty controls_snapshot
is "trivially reconstructible" per ``_residual_reconstructible`` (identity mults,
base == residual).

NOTE ON secondary_loss: ``_make_scenarios`` does not set ``secondary_loss`` on the
Scenario model (it is nullable), so the frozen ``scenario_inputs_snapshot`` carries
``secondary_loss: null``. ``scenario_let_formula`` always accesses ``scen["secondary_
loss"]`` via a direct dict lookup (line 283: ``sl_dist = scen[_K_SL]``), then passes
it to ``_assert_numeric_dist``, which calls ``.items()`` on the resulting value (line
210: ``for key, val in dist_dict.items()``). On a ``None`` secondary_loss the
``.items()`` call in ``_assert_numeric_dist`` raises ``AttributeError``. To get a real
LET we patch ``run.scenario_inputs_snapshot`` so the first scenario entry has a
concrete secondary_loss PERT distribution. This is a test-fixture concern only; it
does not modify any source file.

NOTE ON OPENPYXL AND DYNAMIC-ARRAY FORMULAS: xlsxwriter writes the LET as a
``write_dynamic_array_formula`` call on a single anchor cell. When openpyxl reads the
file, the anchor cell's ``cell.value`` is NOT a plain string — it is an
``openpyxl.worksheet.formula.ArrayFormula`` object. The formula text is in the
``ArrayFormula.text`` attribute and starts with ``=_xlfn.LET(`` (xlsxwriter injects
the ``_xlfn.`` prefix for future functions). The ``_let_formula_strings`` helper
detects the ``ArrayFormula`` type, extracts ``.text``, and checks ``"LET(" in text``
to reliably capture the formula regardless of the ``_xlfn.`` prefix. Plain string
formulas (``cell.value`` is a str) are also checked with the same ``"LET("`` test for
forward-compatibility if openpyxl behaviour changes.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from idraa.services.verification_workbook import _LET_STAT_SPEC, build_verification_workbook
from tests.integration._reports_fixtures import (
    _make_completed_aggregate_run,
    _make_completed_single_run,
)

# A minimal PERT secondary_loss distribution to inject into the fixture snapshot
# so scenario_let_formula receives a non-None secondary_loss dict and emits the LET.
_SL_DIST = {"distribution": "PERT", "low": 500.0, "mode": 5_000.0, "high": 50_000.0}


async def _make_reconstructible_run(db_session, org):
    """Shared builder: a COMPLETED single run that emits a real LET.

    ``controls=[]`` -> trivially reconstructible (identity mults, base==residual).
    The first scenario's ``secondary_loss`` is patched to a concrete PERT so
    ``scenario_let_formula`` (which always accesses ``scen["secondary_loss"]``)
    emits the full LET. See the module docstring for the full rationale.
    """
    run = await _make_completed_single_run(db_session, org, controls=[])
    await db_session.commit()
    sis = run.scenario_inputs_snapshot or {}
    scenarios = list(sis.get("scenarios") or [])
    if scenarios and scenarios[0].get("secondary_loss") is None:
        scenarios[0] = dict(scenarios[0], secondary_loss=_SL_DIST)
        run.scenario_inputs_snapshot = dict(sis, scenarios=scenarios)
    return run


async def _make_reconstructible_aggregate_run(db_session, org):
    """Shared builder: a COMPLETED AGGREGATE run that emits >=2 real LETs.

    Analogous to ``_make_reconstructible_run`` for the aggregate path. The
    run-level ``controls_snapshot`` is empty (``controls=[]``) so every scenario
    is trivially reconstructible (identity mults, base==residual) and
    ``aggregate_control_ids_per_scenario`` is None (engine full-universe fallback)
    so each scenario sees that empty snapshot. ``_make_scenarios`` leaves
    ``secondary_loss`` None, which makes ``scenario_let_formula`` raise on the SL
    dict lookup; so we patch every scenario's ``secondary_loss`` to a concrete PERT
    in the frozen ``scenario_inputs_snapshot``. The result emits one LET per
    scenario => a real multi-scenario roll-up SUM and a live verdict to test.
    """
    run = await _make_completed_aggregate_run(db_session, org, controls=[])
    await db_session.commit()
    # _make_completed_aggregate_run does NOT populate scenario_inputs_snapshot, so
    # build one from the run's frozen per_scenario list (the names/ids the workbook
    # joins on) using the PERT defaults _make_scenarios seeds + a concrete SL PERT.
    per_scenario = (run.simulation_results or {}).get("per_scenario") or []
    scenarios = [
        {
            "scenario_id": str(ps.get("scenario_id")),
            "scenario_name": ps.get("scenario_name"),
            "threat_event_frequency": {
                "distribution": "PERT",
                "low": 0.1,
                "mode": 0.5,
                "high": 1.0,
            },
            "vulnerability": {"distribution": "PERT", "low": 0.2, "mode": 0.4, "high": 0.6},
            "primary_loss": {
                "distribution": "PERT",
                "low": 10_000.0,
                "mode": 100_000.0,
                "high": 1_000_000.0,
            },
            "secondary_loss": _SL_DIST,
        }
        for ps in per_scenario
    ]
    run.scenario_inputs_snapshot = {"scenarios": scenarios}
    return run


async def _make_fail_loud_run(db_session, org):
    """Shared builder: a COMPLETED single run whose controls_snapshot is a LEGACY
    v2 shape, so ``_residual_reconstructible`` returns False and the LET is NOT
    emitted (fail-loud path). The App base figures are still rendered."""
    run = await _make_completed_single_run(
        db_session,
        org,
        controls=[("Firewall", "LOSS_EVENT", "preventive")],  # v2 snapshot -> LegacySnapshotError
    )
    await db_session.commit()
    return run


def _money_values_multiset(ws) -> list[float]:
    vals: list[float] = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                vals.append(round(float(cell.value), 6))
    return sorted(vals)


def _let_formula_strings(ws) -> list[str]:
    """Extract self-contained LET formula strings from a worksheet.

    openpyxl reads xlsxwriter dynamic-array formulas as ArrayFormula objects
    (not plain strings). The formula text lives in ``ArrayFormula.text`` and
    starts with ``=_xlfn.LET(`` (xlsxwriter injects ``_xlfn.`` for future
    functions). This helper handles both the ArrayFormula object shape AND plain
    string formulas (``cell.value`` is a str), so it works if openpyxl behaviour
    changes in a future version.
    """
    from openpyxl.worksheet.formula import ArrayFormula

    out: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            # Dynamic-array formula: openpyxl wraps it in ArrayFormula; text in .text
            if isinstance(v, ArrayFormula):
                text = v.text or ""
                if "LET(" in text:
                    out.append(text)
            # Plain string formula (forward-compat / regular formulas)
            elif isinstance(v, str) and v.startswith("=") and "LET(" in v:
                out.append(v)
    return sorted(out)


# ---------------------------------------------------------------------------
# BASELINE constants — captured from a live run on 2026-06-17 and frozen.
# To update after an intentional FAIR math change: delete these lists, run the
# test with -s to print the new values, then replace and commit with a note.
# ---------------------------------------------------------------------------

BASELINE_MONEY: list[float] = [
    0.0,
    1.0,
    1.0,
    1.0,
    1.0,
    4.0,
    100000.0,
    300000.0,
    400000.0,
    400000.0,
    400000.0,
    400000.0,
    680000.0,
    750000.0,
    800000.0,
    800000.0,
    900000.0,
    1000000.0,
    1200000.0,
    1300000.0,
]

BASELINE_LETS: list[str] = [
    "=_xlfn.LET("
    "_xlpm.u_tef, _xlfn.RANDARRAY(10000,1), "
    "_xlpm.u_vuln, _xlfn.RANDARRAY(10000,1), "
    "_xlpm.u_pl, _xlfn.RANDARRAY(10000,1), "
    "_xlpm.u_sl, _xlfn.RANDARRAY(10000,1), "
    "_xlpm.tef, 0.1 + _xlfn.BETA.INV(_xlpm.u_tef, 3.6808413352, 4.2697759488) * (1.0 - 0.1), "
    "_xlpm.vuln, 0.2 + _xlfn.BETA.INV(_xlpm.u_vuln, 4.0, 4.0) * (0.6 - 0.2), "
    "_xlpm.pl, 10000.0 + _xlfn.BETA.INV(_xlpm.u_pl, 1.2096168295, 4.1126972203) * (1000000.0 - 10000.0), "
    "_xlpm.sl, 500.0 + _xlfn.BETA.INV(_xlpm.u_sl, 1.2096168295, 4.1126972203) * (50000.0 - 500.0), "
    "_xlpm.base_loss, "
    "(_xlpm.tef>0)*_xlpm.tef*((_xlpm.vuln>0)*(_xlpm.vuln<1)*_xlpm.vuln+(_xlpm.vuln>=1))"
    "*((_xlpm.pl>0)*_xlpm.pl+(_xlpm.sl>0)*_xlpm.sl), "
    "_xlpm.tef_r, 0.1 + _xlfn.BETA.INV(_xlpm.u_tef, 3.6808413352, 4.2697759488) * (1.0 - 0.1), "
    "_xlpm.sl_raw, 500.0 + _xlfn.BETA.INV(_xlpm.u_sl, 1.2096168295, 4.1126972203) * (50000.0 - 500.0), "
    "_xlpm.pl_r, 10000.0 + _xlfn.BETA.INV(_xlpm.u_pl, 1.2096168295, 4.1126972203) * (1000000.0 - 10000.0), "
    "_xlpm.vuln_r, (((_xlpm.vuln*1.0)>0)*((_xlpm.vuln*1.0)<1)*(_xlpm.vuln*1.0)+((_xlpm.vuln*1.0)>=1)), "
    "_xlpm.sl_r, (((_xlpm.sl_raw>0)*_xlpm.sl_raw-0.0)>0)*((_xlpm.sl_raw>0)*_xlpm.sl_raw-0.0), "
    "_xlpm.res_loss, "
    "(_xlpm.tef_r>0)*_xlpm.tef_r*_xlpm.vuln_r*((_xlpm.pl_r>0)*_xlpm.pl_r+_xlpm.sl_r), "
    "CHOOSE({1;2;3;4;5;6;7;8;9}, "
    "AVERAGE(_xlpm.base_loss), AVERAGE(_xlpm.res_loss), "
    "AVERAGE(_xlpm.base_loss)-AVERAGE(_xlpm.res_loss), "
    "_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.95), "
    "_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.99), "
    "_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.999), "
    "IFERROR(SUMPRODUCT((_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.95))*_xlpm.res_loss)"
    "/SUMPRODUCT(--(_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.95))), MAX(_xlpm.res_loss)), "
    "IFERROR(SUMPRODUCT((_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.99))*_xlpm.res_loss)"
    "/SUMPRODUCT(--(_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.99))), MAX(_xlpm.res_loss)), "
    "IFERROR(SUMPRODUCT((_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.999))*_xlpm.res_loss)"
    "/SUMPRODUCT(--(_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.999))), MAX(_xlpm.res_loss))))"
]


@pytest.mark.asyncio
async def test_single_run_math_invariants_baseline(db_session, seed_organization_factory):
    org = await seed_organization_factory(name="vwb-formatting-baseline")
    run = await _make_reconstructible_run(db_session, org)

    data = build_verification_workbook(run, org)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["MC"]
    money = _money_values_multiset(ws)
    lets = _let_formula_strings(ws)

    print(f"\nBASELINE_MONEY ({len(money)} values): {money!r}")
    print(f"\nBASELINE_LETS ({len(lets)} formulas):")
    for f in lets:
        print(f"  {f[:120]!r}...")

    assert money == BASELINE_MONEY, (
        f"App numeric values changed: got {len(money)} values, expected {len(BASELINE_MONEY)}"
    )
    assert lets == BASELINE_LETS, (
        f"LET formula strings changed: got {len(lets)}, expected {len(BASELINE_LETS)}"
    )


def test_styles_factory_builds_reusable_formats():
    import xlsxwriter

    from idraa.services.verification_workbook import _Styles

    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    styles = _Styles(wb)
    for key in (
        "title",
        "section_header",
        "money",
        "multiplier",
        "pct",
        "flag_ok",
        "flag_check",
        "note_muted",
        "verdict_ok",
        "verdict_check",
        "header",
        "doc_title",
        "doc_body",
        "doc_heading",
    ):
        assert getattr(styles, key) is not None
    wb.close()


def test_xlsxrows_applies_format_to_cell():
    import xlsxwriter

    from idraa.services.verification_workbook import _Styles, _XlsxRows

    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    ws = wb.add_worksheet("S")
    styles = _Styles(wb)
    rows = _XlsxRows(ws)
    rows.row([("Residual ALE", None), (400000.0, styles.money)])  # (value, fmt) tuples
    wb.close()
    rb = openpyxl.load_workbook(io.BytesIO(buf.getvalue()))["S"]
    assert rb["A1"].value == "Residual ALE"
    assert rb["B1"].value == 400000.0
    assert "$#,##0" in (rb["B1"].number_format or "")


def _banner_text(cell_value) -> str:
    """Coerce a banner cell value (str OR openpyxl ArrayFormula) to its text."""
    from openpyxl.worksheet.formula import ArrayFormula

    if isinstance(cell_value, ArrayFormula):
        return cell_value.text or ""
    return cell_value if isinstance(cell_value, str) else ""


@pytest.mark.asyncio
async def test_workbook_uses_brand_graphite_headers_not_legacy_accent(
    db_session, seed_organization_factory
):
    """Design-system alignment (#59 P3): the produced workbook carries the
    graphite brand (#37464F) header fill and NONE of the legacy ad-hoc accent
    (#E7EEF6) remains. openpyxl reads fills as 8-char ARGB (e.g. 'FF37464F')."""
    org = await seed_organization_factory(name="vwb-brand-navy")
    run = await _make_reconstructible_run(db_session, org)

    wb = openpyxl.load_workbook(io.BytesIO(build_verification_workbook(run, org)))
    fills: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                fg = cell.fill.fgColor
                if fg is not None and fg.rgb and isinstance(fg.rgb, str):
                    fills.add(fg.rgb.upper())
    assert "FF37464F" in fills, f"brand header fill missing; fills={sorted(fills)}"
    assert "FFE7EEF6" not in fills, "legacy _ACCENT #E7EEF6 must be gone"


@pytest.mark.asyncio
async def test_single_run_verdict_first_layout(db_session, seed_organization_factory):
    org = await seed_organization_factory(name="vwb-verdict-layout")
    run = await _make_reconstructible_run(db_session, org)

    wb = openpyxl.load_workbook(io.BytesIO(build_verification_workbook(run, org)))
    ws = wb["MC"]

    assert ws["A1"].value.startswith("VERIFICATION")
    banner = _banner_text(ws["A2"].value)
    assert banner.startswith("=") and "COUNTIF(" in banner, banner
    hdr = [ws.cell(row=4, column=c).value for c in range(1, 7)]
    assert hdr == ["Metric", "Excel (LET)", "App (fair_cam)", "Δ", "Δ%", "ok?"]
    assert "$#,##0" in (ws.cell(row=5, column=3).number_format or "")  # App money
    assert "0.0%" in (ws.cell(row=5, column=5).number_format or "")  # Δ%
    assert ws.freeze_panes == "A5"
    assert ws["A15"].value == "INPUTS"
    # Pin the LOWER edge of the stat region so a future _LET_STAT_SPEC-length
    # desync (flag-range / banner-span) is caught. Row 13 is the last stat's
    # ok? flag; row 14 is the blank gap below the region.
    last_flag = ws.cell(row=13, column=6).value
    assert last_flag is not None and "IF(ABS(" in str(last_flag), last_flag
    assert ws.cell(row=14, column=6).value in (None, "")

    # MECHANISM section: the 9-row LET stat array spills down column B; column A
    # must carry the 9 metric NAMES (one per spill row) instead of a single
    # "stat array (spills down) ->" pointer with 8 anonymous rows below it.
    expected_labels = [spec[0] for spec in _LET_STAT_SPEC]
    label_rows = [
        cell.row
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("LET (metric names in column A")
    ]
    assert label_rows, "MECHANISM LET label row not found"
    anchor_row = label_rows[0] + 1  # LET anchor (and first label row) is just below
    col_a_labels = [
        ws.cell(row=anchor_row + i, column=1).value for i in range(len(expected_labels))
    ]
    assert col_a_labels == expected_labels, col_a_labels
    flat = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)]
    assert "stat array (spills down) ->" not in flat


@pytest.mark.asyncio
async def test_single_run_math_invariants_after_styling(db_session, seed_organization_factory):
    org = await seed_organization_factory(name="vwb-after-styling")
    run = await _make_reconstructible_run(db_session, org)

    ws = openpyxl.load_workbook(io.BytesIO(build_verification_workbook(run, org)))["MC"]
    assert _money_values_multiset(ws) == BASELINE_MONEY
    assert _let_formula_strings(ws) == BASELINE_LETS


# ===========================================================================
# AGGREGATE sheet (Task 3) — verdict-first reorder + formatting, value-preserving
# ===========================================================================

# BASELINE constants — captured from a live run on 2026-06-17 and frozen (CURRENT
# pre-reorder aggregate code). After the reorder these MUST be unchanged; only
# coordinates + formats may move. To re-pin after an intentional FAIR-math change:
# delete, run with -s to print, replace + commit with a note.
# Money multiset (3-scenario reconstructible aggregate): the subset residual-ALE
# roll-up (3 x 400k = 1.2M), the subset base-ALE roll-up (3 x 800k = 2.4M), and the
# App aggregate tail ladder (var_95/99/999 = 3.9M/5.2M/6.8M, es_95/99/999 =
# 4.5M/5.9M/7.4M) from aggregate_with_controls.
BASELINE_AGG_MONEY: list[float] = [
    1_200_000.0,
    2_400_000.0,
    3_900_000.0,
    4_500_000.0,
    5_200_000.0,
    5_900_000.0,
    6_800_000.0,
    7_400_000.0,
]

# Three identical per-scenario LETs (the fixture scenarios share distributions);
# _let_formula_strings returns them sorted, so the list is the same string x3.
_AGG_LET = (
    "=_xlfn.LET("
    "_xlpm.u_tef, _xlfn.RANDARRAY(50000,1), "
    "_xlpm.u_vuln, _xlfn.RANDARRAY(50000,1), "
    "_xlpm.u_pl, _xlfn.RANDARRAY(50000,1), "
    "_xlpm.u_sl, _xlfn.RANDARRAY(50000,1), "
    "_xlpm.tef, 0.1 + _xlfn.BETA.INV(_xlpm.u_tef, 3.6808413352, 4.2697759488) * (1.0 - 0.1), "
    "_xlpm.vuln, 0.2 + _xlfn.BETA.INV(_xlpm.u_vuln, 4.0, 4.0) * (0.6 - 0.2), "
    "_xlpm.pl, 10000.0 + _xlfn.BETA.INV(_xlpm.u_pl, 1.2096168295, 4.1126972203) * (1000000.0 - 10000.0), "
    "_xlpm.sl, 500.0 + _xlfn.BETA.INV(_xlpm.u_sl, 1.2096168295, 4.1126972203) * (50000.0 - 500.0), "
    "_xlpm.base_loss, "
    "(_xlpm.tef>0)*_xlpm.tef*((_xlpm.vuln>0)*(_xlpm.vuln<1)*_xlpm.vuln+(_xlpm.vuln>=1))"
    "*((_xlpm.pl>0)*_xlpm.pl+(_xlpm.sl>0)*_xlpm.sl), "
    "_xlpm.tef_r, 0.1 + _xlfn.BETA.INV(_xlpm.u_tef, 3.6808413352, 4.2697759488) * (1.0 - 0.1), "
    "_xlpm.sl_raw, 500.0 + _xlfn.BETA.INV(_xlpm.u_sl, 1.2096168295, 4.1126972203) * (50000.0 - 500.0), "
    "_xlpm.pl_r, 10000.0 + _xlfn.BETA.INV(_xlpm.u_pl, 1.2096168295, 4.1126972203) * (1000000.0 - 10000.0), "
    "_xlpm.vuln_r, (((_xlpm.vuln*1.0)>0)*((_xlpm.vuln*1.0)<1)*(_xlpm.vuln*1.0)+((_xlpm.vuln*1.0)>=1)), "
    "_xlpm.sl_r, (((_xlpm.sl_raw>0)*_xlpm.sl_raw-0.0)>0)*((_xlpm.sl_raw>0)*_xlpm.sl_raw-0.0), "
    "_xlpm.res_loss, "
    "(_xlpm.tef_r>0)*_xlpm.tef_r*_xlpm.vuln_r*((_xlpm.pl_r>0)*_xlpm.pl_r+_xlpm.sl_r), "
    "CHOOSE({1;2;3;4;5;6;7;8;9}, "
    "AVERAGE(_xlpm.base_loss), AVERAGE(_xlpm.res_loss), "
    "AVERAGE(_xlpm.base_loss)-AVERAGE(_xlpm.res_loss), "
    "_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.95), "
    "_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.99), "
    "_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.999), "
    "IFERROR(SUMPRODUCT((_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.95))*_xlpm.res_loss)"
    "/SUMPRODUCT(--(_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.95))), MAX(_xlpm.res_loss)), "
    "IFERROR(SUMPRODUCT((_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.99))*_xlpm.res_loss)"
    "/SUMPRODUCT(--(_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.99))), MAX(_xlpm.res_loss)), "
    "IFERROR(SUMPRODUCT((_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.999))*_xlpm.res_loss)"
    "/SUMPRODUCT(--(_xlpm.res_loss>=_xlfn.PERCENTILE.INC(_xlpm.res_loss, 0.999))), MAX(_xlpm.res_loss))))"
)
BASELINE_AGG_LETS: list[str] = [_AGG_LET, _AGG_LET, _AGG_LET]


@pytest.mark.asyncio
async def test_aggregate_math_invariants_baseline(db_session, seed_organization_factory):
    org = await seed_organization_factory(name="vwb-agg-baseline")
    run = await _make_reconstructible_aggregate_run(db_session, org)

    data = build_verification_workbook(run, org)
    ws = openpyxl.load_workbook(io.BytesIO(data))["Aggregate"]
    money = _money_values_multiset(ws)
    lets = _let_formula_strings(ws)

    print(f"\nBASELINE_AGG_MONEY ({len(money)} values): {money!r}")
    print(f"\nBASELINE_AGG_LETS ({len(lets)} formulas):")
    for f in lets:
        print(f"  {f[:120]!r}...")

    assert money == BASELINE_AGG_MONEY, (
        f"aggregate App numeric values changed: got {len(money)} values "
        f"({money!r}), expected {len(BASELINE_AGG_MONEY)}"
    )
    assert lets == BASELINE_AGG_LETS, (
        f"aggregate LET formula strings changed: got {len(lets)}, expected {len(BASELINE_AGG_LETS)}"
    )


@pytest.mark.asyncio
async def test_aggregate_verdict_first_layout(db_session, seed_organization_factory):
    org = await seed_organization_factory(name="vwb-agg-verdict-layout")
    run = await _make_reconstructible_aggregate_run(db_session, org)

    ws = openpyxl.load_workbook(io.BytesIO(build_verification_workbook(run, org)))["Aggregate"]

    # Title verdict-first at A1.
    assert ws["A1"].value.startswith("AGGREGATE VERIFICATION"), ws["A1"].value
    # Banner at A2 is a live COUNTIF verdict (>=1 scenario emitted in this fixture).
    banner = _banner_text(ws["A2"].value)
    assert banner.startswith("=") and "COUNTIF(" in banner, banner
    # Header row 4 carries the 6 comparison columns incl. the Δ% column.
    hdr = [ws.cell(row=4, column=c).value for c in range(1, 7)]
    assert hdr == ["Metric", "Excel (LET roll-up)", "App (fair_cam)", "Δ", "Δ%", "ok?"], hdr
    # The first additive-ALE row's App cell (col C, row 5) is money-formatted.
    assert "$#,##0" in (ws.cell(row=5, column=3).number_format or "")
    # Freeze panes set below the header.
    assert ws.freeze_panes == "A5", ws.freeze_panes
    # A PER-SCENARIO section header exists and each emitted scenario has a name band.
    flat = " ".join(str(c.value) for r in ws.iter_rows() for c in r if isinstance(c.value, str))
    assert "PER-SCENARIO MONTE CARLO" in flat, flat[:400]
    for name in ("Ransomware", "Insider", "APT"):
        assert f"Scenario block: {name}" in flat, name
    # Provenance / disclosure relocated below the verdict region (still present).
    assert "in-Excel MC shown for" in flat

    # Each per-scenario block now carries the 9 metric NAMES down column A (one per
    # spill row) instead of a single "stat array (spills down) ->" pointer with 8
    # empty rows beneath it. Locate a block header, then assert the 9 rows below it
    # carry the _LET_STAT_SPEC labels in order in column A.
    expected_labels = [spec[0] for spec in _LET_STAT_SPEC]
    block_rows = [
        cell.row
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value == "Scenario block: Ransomware"
    ]
    assert block_rows, "Ransomware scenario block header not found"
    anchor_row = block_rows[0] + 1  # LET anchor (and label row) is just below header
    col_a_labels = [
        ws.cell(row=anchor_row + i, column=1).value for i in range(len(expected_labels))
    ]
    assert col_a_labels == expected_labels, col_a_labels
    # The old anchor-only pointer text is gone everywhere.
    assert "stat array (spills down) ->" not in flat


@pytest.mark.asyncio
async def test_aggregate_math_invariants_after_styling(db_session, seed_organization_factory):
    org = await seed_organization_factory(name="vwb-agg-after-styling")
    run = await _make_reconstructible_aggregate_run(db_session, org)

    ws = openpyxl.load_workbook(io.BytesIO(build_verification_workbook(run, org)))["Aggregate"]
    assert _money_values_multiset(ws) == BASELINE_AGG_MONEY
    assert _let_formula_strings(ws) == BASELINE_AGG_LETS


@pytest.mark.asyncio
async def test_aggregate_verdict_degenerate_all_excluded(db_session, seed_organization_factory):
    # No reconstructible scenarios (default v2 legacy controls + no scenario_inputs_
    # snapshot) -> zero LETs emitted -> a STATIC amber banner (no live COUNTIF), and
    # every scenario listed summary-only.
    org = await seed_organization_factory(name="vwb-agg-degenerate")
    run = await _make_completed_aggregate_run(db_session, org)  # legacy v2 controls
    await db_session.commit()

    ws = openpyxl.load_workbook(io.BytesIO(build_verification_workbook(run, org)))["Aggregate"]
    assert ws["A1"].value.startswith("AGGREGATE VERIFICATION")
    banner = _banner_text(ws["A2"].value)
    assert "NO RECONSTRUCTIBLE SCENARIOS" in banner, banner
    assert "COUNTIF(" not in banner  # static amber, no live flags
    flat = " ".join(str(c.value) for r in ws.iter_rows() for c in r if isinstance(c.value, str))
    assert "EXCLUDED SCENARIOS (summary-only" in flat
    assert "in-Excel MC shown for 0 of" in flat


@pytest.mark.asyncio
async def test_single_run_verdict_fail_loud(db_session, seed_organization_factory):
    org = await seed_organization_factory(name="vwb-fail-loud")
    run = await _make_fail_loud_run(db_session, org)

    wb = openpyxl.load_workbook(io.BytesIO(build_verification_workbook(run, org)))
    ws = wb["MC"]

    assert ws["A1"].value.startswith("VERIFICATION")
    banner = _banner_text(ws["A2"].value)
    # Static amber banner — no live COUNTIF formula (there are no live flags).
    assert "RESIDUAL NOT RE-DERIVED" in banner
    assert "COUNTIF(" not in banner
    # No green "OK" flags: the ok? column (col 6, rows 5..13) is "n/a".
    flag_vals = [ws.cell(row=r, column=6).value for r in range(5, 14)]
    assert all(v == "n/a" for v in flag_vals), flag_vals
    assert not any(v == "OK" for v in flag_vals)
    # App numbers still render (the base/residual ALE App column is numeric).
    app_vals = [ws.cell(row=r, column=3).value for r in range(5, 14)]
    assert any(isinstance(v, (int, float)) and not isinstance(v, bool) for v in app_vals)


@pytest.mark.asyncio
async def test_documentation_tab_styled(db_session, seed_organization_factory):
    org = await seed_organization_factory(name="vwb-doc-styled")
    run = await _make_reconstructible_run(db_session, org)
    ws = openpyxl.load_workbook(io.BytesIO(build_verification_workbook(run, org)))["Documentation"]
    assert ws["A1"].font.bold is True  # title bold
    assert (ws.column_dimensions["A"].width or 0) >= 60  # column A widened for prose
    assert ws["A1"].alignment.wrap_text is True  # text wrapped
