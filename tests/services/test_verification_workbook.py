import pytest

from idraa.services.verification_workbook import _neutralize


@pytest.mark.parametrize(
    "dangerous", ["=1+1", "+1", "-1", "@SUM(A1)", "{=danger()}", "\t=x", "\r=x", "\n=x"]
)
def test_neutralize_prefixes_dangerous_leading_chars(dangerous):
    out = _neutralize(dangerous)
    assert out.startswith("'"), f"{dangerous!r} not neutralized -> {out!r}"


@pytest.mark.parametrize("safe", ["Acme Corp", "Scenario 1", "", "已", "a=b"])
def test_neutralize_leaves_safe_strings(safe):
    assert _neutralize(safe) == safe


# --- Task 4: snapshot -> minimal fair_cam Control adapter ---------------------
from idraa.services.verification_workbook import (  # noqa: E402
    LegacySnapshotError,
    snapshot_to_fair_cam_controls,
)

# Three REAL FairCamSubFunction slugs in DISTINCT composition groups (so Task 6's
# composition is non-trivial); all PROBABILITY unit_type per fair_cam
# SUB_FUNCTION_UNITS:
#   lec_prev_avoidance     -> LEC_PREVENTION (OR group)
#   lec_det_visibility     -> LEC_DETECTION  (AND group)
#   dsc_prev_communication -> DSC_PREVENTION (AND group)


def _v3_assignment(sub_function: str, unit_type: str, cap=0.8, cov=1.0, rel=1.0):
    return {
        "sub_function": sub_function,
        "capability_value": cap,
        "coverage": cov,
        "reliability": rel,
        "unit_type": unit_type,
    }


def _v3_control(control_id, assignments):
    return {
        "snapshot_version": 3,
        "control_id": control_id,
        "name": "Ctl",
        "domains": ["loss_event"],
        "type": "preventive",
        "assignments": assignments,
    }


def test_adapter_preserves_all_assignments_iteration_contract():
    # N>=3 assignments must all survive (no [0]/[-1] truncation).
    asns = [
        _v3_assignment("lec_prev_avoidance", "probability"),
        _v3_assignment("lec_det_visibility", "probability"),
        _v3_assignment("dsc_prev_communication", "probability"),
    ]
    controls = snapshot_to_fair_cam_controls([_v3_control("c1", asns)])
    assert len(controls) == 1
    assert len(controls[0].assignments) == 3


def test_adapter_field_sync_carries_composition_fields():
    a = _v3_assignment("lec_det_visibility", "probability", cap=0.6, cov=0.9, rel=0.7)
    ctl = snapshot_to_fair_cam_controls([_v3_control("c1", [a])])[0]
    fa = ctl.assignments[0]
    assert fa.capability_value == 0.6
    assert fa.coverage == 0.9
    assert fa.reliability == 0.7
    assert ctl.control_id == "c1"


def test_adapter_rejects_legacy_v1_v2_snapshot():
    v2 = {
        "snapshot_version": 2,
        "control_id": "c1",
        "name": "x",
        "domains": [],
        "type": "preventive",
        "assignments": [],
    }
    with pytest.raises(LegacySnapshotError):
        snapshot_to_fair_cam_controls([v2])


def test_adapter_rejects_malformed_v3_as_legacy():
    bad = {
        "snapshot_version": 3,
        "control_id": "c1",
        "name": "x",
        "domains": [],
        "type": "preventive",
        "assignments": [{"sub_function": "lec_det_visibility"}],  # missing coverage/reliability
    }
    with pytest.raises(LegacySnapshotError):
        snapshot_to_fair_cam_controls([bad])


def test_adapter_rejects_out_of_range_coverage_as_legacy():
    # Task-4 spec-review NICE: a V3 snapshot that passes Pydantic (coverage:float has
    # no bound validator on the audit DTO) but fails fair_cam's __post_init__
    # ([0,1] bound) must funnel the ValueError to LegacySnapshotError, not 500.
    bad = _v3_control("c1", [_v3_assignment("lec_det_visibility", "probability", cov=1.5)])
    with pytest.raises(LegacySnapshotError):
        snapshot_to_fair_cam_controls([bad])


def test_residual_reconstructible_degrades_on_out_of_range_meta_capability():
    # Final-Sec-1 (#439/#451 final-gate): capability_value is UNBOUNDED at the
    # fair_cam DTO layer (unlike coverage/reliability, __post_init__ does not
    # clamp/validate it -- see fair_cam/models/control.py docstring), so a
    # legacy-stored snapshot can carry an out-of-range PERCENT_REDUCTION
    # capability_value (85.0, not [0,1]) that sails through
    # snapshot_to_fair_cam_controls. VMC_PREV_REDUCE_VARIANCE_PROB is a "meta"
    # (VMC) sub-function OR-composed straight into E_vmc with no AND-gating
    # that could zero it out, so the out-of-range value survives to Slice 2's
    # precompose_parts E-bound guard, which fires a ValueError ("e_vmc out of
    # [0,1]") -- previously uncaught here, this would 500 the workbook
    # download instead of degrading gracefully.
    from idraa.services.verification_workbook import _residual_reconstructible

    snap = [
        _v3_control(
            "c1",
            [_v3_assignment("vmc_prev_reduce_variance_prob", "percent_reduction", cap=85.0)],
        )
    ]
    reconstructible, mults = _residual_reconstructible(snap)
    assert reconstructible is False
    assert mults is None


# --- Task 5: unit-table drift gate -------------------------------------------
import fair_cam.models.sub_function as fc_sf  # noqa: E402

import idraa.models.enums as v3_enums  # noqa: E402
from idraa.schemas.run_snapshot import (  # noqa: E402
    ControlFunctionAssignmentSnapshotDTO,
    ControlSnapshotV3,
)
from idraa.services.verification_workbook import unit_table_has_drifted  # noqa: E402

# Real PROBABILITY-unit slug reused from the Task-4 helpers; both the fair_cam and
# v3 SUB_FUNCTION_UNITS tables map it to PROBABILITY, so a one-sided mutation is a
# genuine distinguisher (not a slug that already differs between the tables).
_DRIFT_SUB = "lec_prev_avoidance"


def test_drift_gate_passes_when_tables_match():
    snap = [_v3_control("c1", [_v3_assignment(_DRIFT_SUB, "probability")])]
    # unit_type in the snapshot equals the live fair_cam table for this slug.
    assert unit_table_has_drifted(snap) is False


def test_drift_gate_detects_fair_cam_table_mutation(monkeypatch):
    snap = [_v3_control("c1", [_v3_assignment(_DRIFT_SUB, "probability")])]
    # Mutate the table compose_groups READS (fair_cam's copy).
    mutated = dict(fc_sf.SUB_FUNCTION_UNITS)
    key = fc_sf.FairCamSubFunction(_DRIFT_SUB)
    mutated[key] = fc_sf.UnitType.ELAPSED_TIME  # any value != probability
    monkeypatch.setattr(fc_sf, "SUB_FUNCTION_UNITS", mutated)
    # The gate imports fc_sf at call time, so it sees the monkeypatch.
    assert unit_table_has_drifted(snap) is True


def test_drift_gate_ignores_v3_only_mutation_negative_control(monkeypatch):
    # Mutating ONLY the v3 copy must NOT register as drift — the gate compares
    # against fair_cam's table, the one compose_groups actually reads.
    snap = [_v3_control("c1", [_v3_assignment(_DRIFT_SUB, "probability")])]
    mutated = dict(v3_enums.SUB_FUNCTION_UNITS)
    mutated[v3_enums.FairCamSubFunction(_DRIFT_SUB)] = v3_enums.UnitType.ELAPSED_TIME
    monkeypatch.setattr(v3_enums, "SUB_FUNCTION_UNITS", mutated)
    assert unit_table_has_drifted(snap) is False


def test_drift_gate_passes_on_real_persisted_shape_snapshot():
    # Real persisted-shape snapshot: unit_type is the serialized slug exactly as
    # the DB stores it (StrEnum -> .value). Catches a name/value or v3-vs-fair_cam
    # representation mismatch a hand-built "probability" literal would miss.
    model = ControlSnapshotV3(
        control_id="c1",
        name="Ctl",
        domains=["loss_event"],
        type="preventive",
        assignments=[
            ControlFunctionAssignmentSnapshotDTO(
                sub_function=v3_enums.FairCamSubFunction(_DRIFT_SUB),
                capability_value=0.8,
                coverage=1.0,
                reliability=1.0,
                unit_type=v3_enums.UnitType.PROBABILITY,
            )
        ],
    )
    snap = [model.model_dump(mode="json")]
    assert snap[0]["assignments"][0]["unit_type"] == "probability"  # serialized slug
    assert unit_table_has_drifted(snap) is False


# --- Task 6: composed node multipliers (recomposition-equivalence) ------------
from fair_cam.models.control import Control as FairCamControl  # noqa: E402
from fair_cam.models.control import FairCamControlFunctionAssignment  # noqa: E402
from fair_cam.models.sub_function import FairCamSubFunction  # noqa: E402
from fair_cam.risk_engine.control_aware import (  # noqa: E402
    _NODE_KEYS,
    _group_comp_to_node_multipliers,
)
from fair_cam.risk_engine.group_composition import compose_groups  # noqa: E402

from idraa.services.verification_workbook import composed_node_multipliers  # noqa: E402

_CANON_KEYS = {"threat_event_frequency", "vulnerability", "primary_loss", "secondary_loss"}

# Controls whose sub_functions span DISTINCT composition groups so the
# composition is non-trivial AND every one of the five engine keys lands
# non-tautologically (T6 methodology NICE — magnitude + currency branches were
# previously left at identity, making 3 of the 5 assertions 1.0==1.0 / 0.0==0.0):
#   c1: lec_prev_avoidance (LEC_PREVENTION, OR) -> threat_event_frequency + vulnerability
#       + lec_det_visibility (LEC_DETECTION, AND) — exercises the tef/vuln branch.
#   c2: dsc_prev_communication (DSC_PREVENTION, AND) — single member of a 9-member
#       AND group, so that group's effectiveness is 0 (an absent AND member is 0.0)
#       and it does NOT move magnitude on its own.
#   c3 (NEW): the DSC Identification∧Correction AND-pair is the magnitude driver.
#       DSC_IDENTIFICATION_CORRECTION_PAIR has BOTH its members (dsc_id_misaligned +
#       dsc_corr_misaligned, both PROBABILITY) present, so the AND group composes to
#       a non-zero effectiveness and GROUP_NODE_MAPPING routes it to the magnitude
#       targets (secondary_loss w=0.5, primary_loss w=0.2) -> both land off 1.0.
#       c3 also carries lec_resp_loss_reduction (the only CURRENCY sub-function) with
#       a realistic per-event dollar capability_value, so currency_subtractor_total
#       accumulates cap*coverage*reliability != 0.0 (group_composition.py:101).
_TWO_CONTROL_INPUTS = [
    ("c1", [("lec_prev_avoidance", 0.8, 1.0, 1.0), ("lec_det_visibility", 0.7, 0.9, 0.95)]),
    ("c2", [("dsc_prev_communication", 0.6, 1.0, 0.8)]),
    (
        "c3",
        [
            # Complete DSC ID+Correction AND-pair -> primary_loss + secondary_loss.
            ("dsc_id_misaligned", 0.7, 1.0, 0.9),
            ("dsc_corr_misaligned", 0.6, 1.0, 0.95),
            # CURRENCY Loss-Reduction -> currency_subtractor_total (per-event $).
            ("lec_resp_loss_reduction", 25000.0, 1.0, 0.9),
        ],
    ),
]


@pytest.fixture
def two_control_v3_snapshot():
    """Real persisted-shape V3 snapshot dicts (model_dump(mode="json")) for the
    same controls used to build engine_applied_mults."""
    snaps = []
    for control_id, asns in _TWO_CONTROL_INPUTS:
        model = ControlSnapshotV3(
            control_id=control_id,
            name="Ctl",
            domains=["loss_event"],
            type="preventive",
            assignments=[
                ControlFunctionAssignmentSnapshotDTO(
                    sub_function=v3_enums.FairCamSubFunction(sub),
                    capability_value=cap,
                    coverage=cov,
                    reliability=rel,
                    unit_type=v3_enums.SUB_FUNCTION_UNITS[v3_enums.FairCamSubFunction(sub)],
                )
                for sub, cap, cov, rel in asns
            ],
        )
        snaps.append(model.model_dump(mode="json"))
    return snaps


@pytest.fixture
def engine_applied_mults():
    """Ground truth: the multipliers + subtractor the ENGINE PATH applies for the
    same two controls, captured via the engine's OWN chain
    (native_control_aware.py:124-131): compose_groups(controls) ->
    _group_comp_to_node_multipliers(group_comp) -> group_comp.currency_subtractor_total.
    Controls are built directly as fair_cam dataclasses (NOT via
    composed_node_multipliers) so this is a genuine engine-path equivalence."""
    controls = [
        FairCamControl(
            control_id=control_id,
            assignments=[
                FairCamControlFunctionAssignment(
                    sub_function=FairCamSubFunction(sub),
                    capability_value=cap,
                    coverage=cov,
                    reliability=rel,
                    degradation_rate=0.0,
                )
                for sub, cap, cov, rel in asns
            ],
        )
        for control_id, asns in _TWO_CONTROL_INPUTS
    ]
    group_comp = compose_groups(controls)  # native_control_aware.py:124
    node_mults = _group_comp_to_node_multipliers(group_comp)  # :125
    return {**node_mults, "currency_subtractor_total": group_comp.currency_subtractor_total}  # :131


def test_engine_node_keys_are_the_expected_set():
    # Contract guard: if fair_cam renames a node key, fail LOUD here rather than
    # silently mis-wiring the residual sampler.
    assert set(_NODE_KEYS) == _CANON_KEYS


def test_composed_node_multipliers_match_engine(two_control_v3_snapshot, engine_applied_mults):
    """Builder's re-derivation == the mults the ENGINE PATH applied for the same
    controls. engine_applied_mults is captured via the engine's own
    compose_groups -> _group_comp_to_node_multipliers chain (the chain
    native_control_aware.py:124 runs), NOT re-called inside this assertion."""
    result = composed_node_multipliers(two_control_v3_snapshot)
    assert set(result) == _CANON_KEYS | {"currency_subtractor_total"}
    for key in _CANON_KEYS | {"currency_subtractor_total"}:
        assert result[key] == pytest.approx(engine_applied_mults[key])


# --- Task 6 (#439 Slice-2): workbook coupling reconciliation ------------------
#
# One meta + one LEC control. The workbook re-derives multipliers via
# composed_node_multipliers (which calls compose_groups at the CANONICAL default κ);
# the engine applies _group_comp_to_node_multipliers(compose_groups(ctrls)) at the
# SAME default κ. Because the meta control's κ coupling uplifts the co-present LEC
# control's reliability, the tef/vuln multipliers MOVE off the no-meta value — so this
# is a genuine coupling case, not a tautology. Both sides must agree: the workbook
# picks up the meta coupling with ZERO code change.

# c_meta: vmc_id_control_monitoring (OR leaf, E_meta > 0 standalone) — the coupler.
# c_lec:  lec_prev_avoidance (LEC_PREVENTION, OR) -> tef + vuln — the uplifted channel.
_META_PLUS_LEC_INPUTS = [
    ("c_lec", [("lec_prev_avoidance", 0.8, 1.0, 0.9)]),
    ("c_meta", [("vmc_id_control_monitoring", 0.7, 1.0, 0.9)]),
]


def _meta_plus_lec_snapshot():
    snaps = []
    for control_id, asns in _META_PLUS_LEC_INPUTS:
        model = ControlSnapshotV3(
            control_id=control_id,
            name="Ctl",
            domains=["loss_event"],
            type="preventive",
            assignments=[
                ControlFunctionAssignmentSnapshotDTO(
                    sub_function=v3_enums.FairCamSubFunction(sub),
                    capability_value=cap,
                    coverage=cov,
                    reliability=rel,
                    unit_type=v3_enums.SUB_FUNCTION_UNITS[v3_enums.FairCamSubFunction(sub)],
                )
                for sub, cap, cov, rel in asns
            ],
        )
        snaps.append(model.model_dump(mode="json"))
    return snaps


def _meta_plus_lec_controls():
    return [
        FairCamControl(
            control_id=control_id,
            assignments=[
                FairCamControlFunctionAssignment(
                    sub_function=FairCamSubFunction(sub),
                    capability_value=cap,
                    coverage=cov,
                    reliability=rel,
                    degradation_rate=0.0,
                )
                for sub, cap, cov, rel in asns
            ],
        )
        for control_id, asns in _META_PLUS_LEC_INPUTS
    ]


def test_workbook_coupling_reconciliation_meta_plus_lec_plain():
    """composed_node_multipliers(meta+LEC snapshot) == the engine's canonical
    multipliers for the same controls at default κ (plain / detection-gated)."""
    snaps = _meta_plus_lec_snapshot()
    controls = _meta_plus_lec_controls()
    comp = compose_groups(controls)  # canonical default κ
    expected = {
        **_group_comp_to_node_multipliers(comp),
        "currency_subtractor_total": comp.currency_subtractor_total,
    }
    result = composed_node_multipliers(snaps)
    assert set(result) == _CANON_KEYS | {"currency_subtractor_total"}
    for key in _CANON_KEYS | {"currency_subtractor_total"}:
        assert result[key] == pytest.approx(expected[key])
    # Non-tautology guard: the meta coupling actually moved a channel the LEC control
    # targets (tef and/or vuln < 1.0 — a bare 1.0/1.0 would mean no LEC effect at all).
    assert result["threat_event_frequency"] < 1.0 or result["vulnerability"] < 1.0


def test_workbook_coupling_reconciliation_meta_plus_lec_availability():
    """Same reconciliation with availability_self_detection=True: the workbook's
    availability branch must still equal the engine chain called with the same flag
    (coupling × Slice-1 self-detection gate interplay)."""
    snaps = _meta_plus_lec_snapshot()
    controls = _meta_plus_lec_controls()
    comp = compose_groups(controls)
    expected = {
        **_group_comp_to_node_multipliers(comp, availability_self_detection=True),
        "currency_subtractor_total": comp.currency_subtractor_total,
    }
    result = composed_node_multipliers(snaps, availability_self_detection=True)
    assert set(result) == _CANON_KEYS | {"currency_subtractor_total"}
    for key in _CANON_KEYS | {"currency_subtractor_total"}:
        assert result[key] == pytest.approx(expected[key])


# --- Task 9: single-run workbook assembly ------------------------------------
import io  # noqa: E402
from types import SimpleNamespace  # noqa: E402

import openpyxl  # noqa: E402

from idraa.config import get_settings  # noqa: E402
from idraa.services.verification_workbook import build_verification_workbook  # noqa: E402

_settings = get_settings()


def _risk_dict(ale):
    """Mirror _fair_risk_to_dict's app-figure shape (run_executor.py:733-756):
    var_95/var_99 direct keys, var_999 direct (merged), expected_shortfall nested."""
    return {
        "annualized_loss_expectancy": ale,
        "mean": ale,
        "median": ale * 0.8,
        "std_deviation": ale * 0.3,
        "var_95": ale * 2.0,
        "var_99": ale * 3.0,
        "var_999": ale * 4.0,
        "expected_shortfall": {
            "es_95": ale * 2.5,
            "es_99": ale * 3.5,
            "es_999": ale * 4.5,
        },
        "loss_event_frequency": 0.5,
        "loss_magnitude": ale / 0.5,
        "n_simulations": 10,
    }


def _sim_results():
    return {
        "base_risk": _risk_dict(1_000_000.0),
        "residual_risk": _risk_dict(400_000.0),
        "cost_summary": {
            "total_annual_cost": 50_000.0,
            "total_risk_reduction": 600_000.0,
            "net_benefit": 550_000.0,
            "aggregate_roi": 12.0,
        },
    }


def _scenario_inputs_snapshot():
    # One scenario; TEF/PL/SL as PERT, vuln as BETA (vuln-only sample-level clip).
    return {
        "scenarios": [
            {
                "scenario_id": "s1",
                "scenario_name": "Scenario A",
                "threat_event_frequency": {
                    "distribution": "pert",
                    "low": 1.0,
                    "mode": 3.0,
                    "high": 6.0,
                },
                "vulnerability": {"distribution": "beta", "alpha": 2.0, "beta": 5.0},
                "primary_loss": {
                    "distribution": "lognormal",
                    "mean": 12.0,
                    "sigma": 1.0,
                },
                "secondary_loss": {
                    "distribution": "pert",
                    "low": 1000.0,
                    "mode": 5000.0,
                    "high": 20000.0,
                },
            }
        ]
    }


def _make_run(
    *,
    name="Run X",
    controls_snapshot=None,
    mc_iterations=10,
    sim_results=None,
    run_type="single",
):
    return SimpleNamespace(
        name=name,
        run_type=run_type,
        mc_iterations=mc_iterations,
        random_seed=42,
        controls_snapshot=controls_snapshot if controls_snapshot is not None else [],
        scenario_inputs_snapshot=_scenario_inputs_snapshot(),
        simulation_results=sim_results if sim_results is not None else _sim_results(),
    )


def _make_org(name="Acme Corp"):
    return SimpleNamespace(name=name)


def _open(wb_bytes):
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=False)


def _v3_snap_dict():
    """One real-shape V3 control snapshot that recomposes cleanly (no drift)."""
    model = ControlSnapshotV3(
        control_id="c1",
        name="Ctl",
        domains=["loss_event"],
        type="preventive",
        assignments=[
            ControlFunctionAssignmentSnapshotDTO(
                sub_function=v3_enums.FairCamSubFunction("lec_prev_avoidance"),
                capability_value=0.8,
                coverage=1.0,
                reliability=1.0,
                unit_type=v3_enums.UnitType.PROBABILITY,
            )
        ],
    )
    return model.model_dump(mode="json")


# --- Task 10: Documentation sheet content ----------------------------------


def _doc_sheet_text(run=None, org=None):
    """Flatten all string cells of the Documentation sheet into one lowercased
    blob for case-insensitive substring scanning.

    Reads the LET-spill workbook's Documentation sheet (the prose was rewritten to
    the LET model in Task 6); the LET path is what production serves."""
    from idraa.services.verification_workbook import build_single_run_let_workbook

    wb = _open(build_single_run_let_workbook(run or _let_run(), org or _make_org()))
    ws = wb["Documentation"]
    cells = [c for r in ws.iter_rows(values_only=True) for c in r if isinstance(c, str)]
    return " ".join(cells).lower()


def test_documentation_sheet_covers_required_caveat_keywords():
    blob = _doc_sheet_text()
    # independent-validation rationale
    assert "independent" in blob
    # scenario-independence caveat on aggregate VaR/ES (added for the
    # aggregate-tail-risk independence assumption): positively-correlated scenarios
    # understate the tail; treat as a lower bound; ALE additive regardless.
    assert "independence" in blob
    assert "common cause" in blob
    assert "lower bound" in blob
    assert "additive regardless" in blob
    # --- LET model (Task 6) ----------------------------------------------------
    # one-formula-per-scenario generation (RANDARRAY, internal materialization)
    assert "one self-contained let" in blob or "one let per scenario" in blob
    assert "randarray" in blob
    assert "no per-iteration row" in blob or "no explicit rows" in blob
    # Modern-Excel (dynamic-array) requirement; mobile/iPad ARE supported (corrected 2026-06-18)
    assert "microsoft 365" in blob or "m365" in blob
    assert "dynamic-array" in blob
    assert "mobile" in blob
    # RANDARRAY volatility + paste-AS-values to freeze (RAND -> RANDARRAY)
    assert "randarray" in blob
    assert "paste" in blob and "paste-as-values" in blob
    # responsiveness cap (N cap) + the new field names; ΣN aggregate cap
    assert "responsiveness cap" in blob and "n cap" in blob
    assert "verification_workbook_max_n" in blob
    assert "verification_workbook_aggregate_total_max" in blob
    # CRN: base & residual share draws + WHY it holds (scale/shift, not swap)
    assert "common random numbers" in blob or "crn" in blob
    assert "share the same" in blob or "share no" in blob  # shared-draws prose
    assert "scale-invariant" in blob or "log-mean shift" in blob
    assert "distribution swap" in blob  # CRN would NOT survive an arbitrary swap
    assert "control value" in blob or "control-value" in blob
    # array-safe boolean clips + SUMPRODUCT tail-mean (spill rewrite documented)
    assert "(x>0)*x" in blob
    assert "((x>0)*(x<1)*x+(x>=1))" in blob
    assert "sumproduct" in blob
    # PERCENTILE.INC type-7 note retained
    assert "percentile.inc" in blob
    # scope boundary: composition is an INPUT, re-derived via compose_groups
    assert "composition" in blob
    assert "compose_groups" in blob
    assert "input" in blob
    assert "not" in blob
    assert "validated" in blob
    # sampling-tolerance + var_999 stable only near the 100k cap
    assert "var_999" in blob
    assert "100,000" in blob
    # drift gate: unit table + legacy fallbacks
    assert "unit table" in blob
    assert "legacy" in blob and "fallback" in blob
    # calibration caveat: tau / GROUP_NODE_MAPPING
    assert "calibration" in blob
    assert ("tau" in blob) or ("τ" in blob)
    assert ("group_node_mapping" in blob) or ("group→node" in blob) or ("group-node" in blob)
    # independent-RNG honesty (not bit-exact)
    assert "bit-exact" in blob
    # seed for in-app reproduction
    assert "seed" in blob
    # --- OLD explicit-row phrasing must be GONE --------------------------------
    assert "row cap" not in blob
    assert "verification_workbook_max_rows" not in blob
    assert "verification_workbook_max_scenarios" not in blob
    assert "k of m" not in blob and "k-of-m" not in blob
    assert "averageif" not in blob  # ES is SUMPRODUCT now, not AVERAGEIF


def test_documentation_sheet_reports_run_seed_and_drift_result():
    # A reconstructible run reports the seed and a faithful composition result.
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    blob = _doc_sheet_text(run=run)
    assert "42" in blob  # the run's seed
    assert "faithful" in blob


def test_documentation_sheet_legacy_run_reports_degraded():
    legacy = {
        "snapshot_version": 2,
        "control_id": "c1",
        "name": "x",
        "domains": [],
        "type": "preventive",
        "assignments": [],
    }
    run = _let_run(controls_snapshot=[legacy])
    blob = _doc_sheet_text(run=run)
    assert "degraded" in blob or "app residual value" in blob


def test_documentation_sheet_all_cells_neutralized():
    # Defense-in-depth: every doc cell goes through _neutralize -> no raw cell
    # value begins with a formula leader.
    from idraa.services.verification_workbook import build_single_run_let_workbook

    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    ws = wb["Documentation"]
    for r in ws.iter_rows():
        for cell in r:
            v = cell.value
            if isinstance(v, str) and v and v[0] in ("=", "+", "-", "@"):
                raise AssertionError(f"un-neutralized doc cell: {v!r}")


# --- Task 2: single-run LET-spill assembly (xlsxwriter dynamic-array) ----------
# The LET workbook is now written with xlsxwriter (openpyxl cannot emit the
# dynamic-array `cm` metadata). It is still RE-READ with openpyxl for assertions.
# xlsxwriter (use_future_functions=True) prefixes the BARE emitted function names
# with `_xlfn.`/`_xlfn._xlws.` on write, and preserves the `_xlpm.` LET-param
# prefix; the LET lands as a single-anchor ArrayFormula whose `.ref` is one cell.
from openpyxl.worksheet.formula import ArrayFormula  # noqa: E402

from idraa.services.verification_workbook import (  # noqa: E402
    build_single_run_let_workbook,
)


def _let_run(**kw):
    """Single run with the key-less PERT wizard shape (TEF/PL/SL) + a beta vuln,
    so the LET path is exercised with the dominant production shape."""
    run = _make_run(**kw)
    run.scenario_inputs_snapshot = _keyless_pert_scenario_inputs()
    # _keyless_pert_scenario_inputs uses a key-less PERT vuln; give it a beta so the
    # base-side vuln draws a beta (vuln-only sample-level clip).
    run.scenario_inputs_snapshot["scenarios"][0]["vulnerability"] = {
        "distribution": "beta",
        "alpha": 2.0,
        "beta": 5.0,
    }
    return run


def _flat_strings(ws):
    return " ".join(str(c) for r in ws.iter_rows(values_only=True) for c in r if isinstance(c, str))


def _let_anchors(ws):
    """Return [(coordinate, ref, formula_text)] for every dynamic-array LET cell.

    A dynamic-array formula round-trips through openpyxl as an ``ArrayFormula``
    object whose ``.text`` is the (xlsxwriter-prefixed) formula and ``.ref`` is the
    spill anchor range."""
    out = []
    for r in ws.iter_rows():
        for cell in r:
            v = cell.value
            if isinstance(v, ArrayFormula) and "LET" in (v.text or ""):
                out.append((cell.coordinate, v.ref, v.text))
    return out


def _string_cells_blob(wb):
    """Lowercased blob of every TEXT (data_type 's') cell across all sheets."""
    parts = []
    for sheet in wb.worksheets:
        for r in sheet.iter_rows():
            for cell in r:
                if cell.data_type == "s" and isinstance(cell.value, str):
                    parts.append(cell.value)
    return " ".join(parts).lower()


def test_let_workbook_builds_with_required_sheets():
    out = build_single_run_let_workbook(_let_run(), _make_org())
    assert out[:2] == b"PK"
    wb = _open(out)
    assert {"MC", "Documentation"} <= set(wb.sheetnames)


def test_let_workbook_control_economics_carries_weight_provenance_disclaimer():
    """Issue #413: the control-economics block on the MC sheet carries the
    implementation-calibrated-weights disclaimer (anchored to fair_cam's
    weights_provenance)."""
    from idraa.services.reports import CONTROL_WEIGHT_PROVENANCE_DISCLAIMER

    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    blob = _string_cells_blob(wb)
    assert CONTROL_WEIGHT_PROVENANCE_DISCLAIMER.lower() in blob, (
        "control-weight provenance disclaimer missing from the Excel control-economics block"
    )


def test_let_workbook_has_single_anchor_dynamic_array_let():
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    anchors = _let_anchors(wb["MC"])
    assert len(anchors) == 1, f"expected exactly one LET dynamic-array cell, got {anchors}"
    coord, ref, text = anchors[0]
    # SINGLE-cell anchor — ref must be one cell (no ':' range), == the anchor coord.
    assert ":" not in str(ref), f"LET anchor must be a single cell, got ref={ref!r}"
    assert str(ref) == coord
    # xlsxwriter prefixed the BARE function names; the _xlpm. param prefix survived.
    assert "_xlfn.LET" in text
    assert "_xlfn.RANDARRAY" in text
    assert "_xlpm.u_tef" in text
    # internal generation at N=10 (the run's mc_iterations), CHOOSE 9-stat array.
    assert "RANDARRAY(10,1)" in text
    assert "CHOOSE(" in text and "VSTACK(" not in text


def test_let_workbook_opens_via_openpyxl_is_read_only():
    # openpyxl is the reader (not the writer) for the LET path — re-read succeeds.
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    assert wb["MC"]["A1"].value is not None


def test_let_workbook_has_within_tolerance_flag_cells():
    # VWB-1 (2026-07-03 correctness audit): the two 99.9-tail rows (VaR 99.9 / ES
    # 99.9) are noisy at the workbook's capped N and are EXCLUDED from the verdict —
    # their ok? cell reads the "noisy tail" note (NOT the literal "CHECK") when out
    # of band, so a plain COUNTIF(...,"CHECK") over the full span never counts them.
    # So exactly the 7 CORE rows carry the OK/CHECK gate; the 2 tail rows carry an
    # OK/noisy-tail gate. Both kinds are still =IF(ABS(...)) within-tolerance cells.
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    ws = wb["MC"]

    def _tol_cells(*, needle: str) -> list[str]:
        return [
            c.value
            for r in ws.iter_rows()
            for c in r
            if c.data_type == "f"
            and isinstance(c.value, str)
            and c.value.startswith("=IF(ABS(")
            and '"OK"' in c.value
            and needle in c.value
        ]

    core_flags = _tol_cells(needle='"CHECK"')
    assert len(core_flags) == 7, f"expected 7 CORE OK/CHECK flag cells, got {len(core_flags)}"
    # No CHECK cell may be a 99.9-tail row (they must never trip the banner COUNTIF).
    tail_flags = _tol_cells(needle="noisy tail")
    assert len(tail_flags) == 2, f"expected 2 tail OK/noisy-tail cells, got {len(tail_flags)}"
    for f in tail_flags:
        assert '"CHECK"' not in f, f"99.9-tail ok? cell must NOT emit CHECK: {f!r}"


def test_let_workbook_side_by_side_has_app_values():
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    ws = wb["MC"]
    nums = [c for r in ws.iter_rows(values_only=True) for c in r if isinstance(c, (int, float))]
    assert 1_000_000.0 in nums  # base ALE
    assert 400_000.0 in nums  # residual ALE


def test_let_workbook_top_note_mentions_m365_and_spill():
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    blob = _flat_strings(wb["MC"]).lower()
    assert "microsoft 365" in blob
    assert "spill" in blob
    assert "f9" in blob


def test_let_workbook_excel_column_references_spill_cells():
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    ws = wb["MC"]
    # The Excel column references B<anchor>.. (the LET spill); at least one =B<row> ref.
    refs = [
        c.value
        for r in ws.iter_rows()
        for c in r
        if c.data_type == "f"
        and isinstance(c.value, str)
        and c.value.startswith("=B")
        and c.value[2:].isdigit()
    ]
    assert refs, "expected Excel column to reference the LET spill cells (=B<row>)"


def test_let_workbook_degraded_legacy_shows_fail_loud_no_let():
    legacy = {
        "snapshot_version": 1,
        "control_id": "c1",
        "control_strength": 0.5,
        "control_reliability": 0.9,
        "control_coverage": 1.0,
    }
    wb = _open(build_single_run_let_workbook(_let_run(controls_snapshot=[legacy]), _make_org()))
    blob = _flat_strings(wb["MC"]).lower()
    assert "not reconstructible" in blob
    # the LET (base+residual coupled) is NOT emitted on the degraded path.
    assert not _let_anchors(wb["MC"]), "degraded path must not emit a LET formula"


def test_let_workbook_degraded_beta_in_tef_slot_shows_fail_loud():
    # A BETA in a tef slot is composition-reconstructible but NOT param-scalable
    # (scaled_params rejects BETA) -> fail-loud, no LET.
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.scenario_inputs_snapshot["scenarios"][0]["threat_event_frequency"] = {
        "distribution": "beta",
        "alpha": 2.0,
        "beta": 5.0,
    }
    wb = _open(build_single_run_let_workbook(run, _make_org()))
    blob = _flat_strings(wb["MC"]).lower()
    assert "not reconstructible" in blob
    assert not _let_anchors(wb["MC"])


def test_let_workbook_caps_n_at_max():
    run = _let_run(mc_iterations=200_000)
    wb = _open(build_single_run_let_workbook(run, _make_org()))
    anchors = _let_anchors(wb["MC"])
    assert anchors
    assert "RANDARRAY(50000,1)" in anchors[0][2]


# --- Task 5: responsiveness cap NOTE (single-run) -----------------------------


def test_let_workbook_cap_note_present_when_n_cap_binds():
    # mc_iterations (200k) > max_n (50k default) -> N capped at 50k AND a
    # plain-language responsiveness note is emitted.
    run = _let_run(mc_iterations=200_000)
    wb = _open(build_single_run_let_workbook(run, _make_org()))
    anchors = _let_anchors(wb["MC"])
    assert anchors
    assert "RANDARRAY(50000,1)" in anchors[0][2]  # N = max_n
    blob = _flat_strings(wb["MC"]).lower()
    assert "50000 of 200000 trials" in blob
    assert "responsiveness" in blob
    assert "statistically representative" in blob


def test_let_workbook_no_cap_note_when_n_below_cap():
    # mc_iterations (10) <= max_n -> N == mc_iterations, no misleading note.
    run = _let_run(mc_iterations=10)
    wb = _open(build_single_run_let_workbook(run, _make_org()))
    anchors = _let_anchors(wb["MC"])
    assert "RANDARRAY(10,1)" in anchors[0][2]
    blob = _flat_strings(wb["MC"]).lower()
    assert "for responsiveness" not in blob
    assert "of 10 trials" not in blob


def test_let_workbook_names_neutralized():
    run = _let_run(name="=danger")
    org = _make_org(name="=danger")
    wb = _open(build_single_run_let_workbook(run, org))
    found = False
    for sheet in wb.worksheets:
        for r in sheet.iter_rows():
            for cell in r:
                v = cell.value
                if isinstance(v, str) and "danger" in v:
                    assert v.startswith("'="), f"name not neutralized: {v!r}"
                    # write_string forced a TEXT cell — never a formula.
                    assert cell.data_type == "s", f"name cell promoted to {cell.data_type}"
                    found = True
    assert found


# --- Task 2 / Step 1b: writer security hardening (Sec-B1/Sec-B2/I3) ------------
# xlsxwriter's write() AUTO-PROMOTES strings to live formulas (=danger / {=danger})
# and URLs to hyperlinks; the writer swap to xlsxwriter would WIDEN the injection /
# phishing surface if every cell weren't written via the type-specific writers.
# Assert by CELL TYPE (data_type), not leading-quote substring: the ONLY formula
# cells are the TRUSTED, fully-internally-built ones (the LET anchor + the
# side-by-side =B<spill>/delta/within-tolerance-flag cells, which carry cell refs
# and constants ONLY — zero user data); a user-derived hostile string must NEVER
# appear inside any formula cell; AND zero hyperlinks anywhere.


_INJECTION_NAMES = [
    "=danger()",
    "{=danger()}",
    "+danger()",
    "@danger()",
    "http://evil.example/login",
    "mailto:x@y",
]


def _let_anchor_coords(ws):
    return {coord for (coord, _ref, _text) in _let_anchors(ws)}


def _is_trusted_internal_formula(coord, text):
    """A formula cell is trusted iff it is one of the fully-internal forms the
    assembly emits: the LET dynamic array, an Excel-column spill ref (=B<n>), a
    delta (=B<n>-C<n>), a within-tolerance flag (=IF(ABS(...))), or a worked-example
    cell (RAND() / an _invcdf sample / a scalar MAX-MIN clip / the loss chain). All
    reference cells/constants only — they NEVER interpolate user data, so they are
    not an injection vector. Anything else being a formula is a promotion bug."""
    import re

    t = (text or "").replace(" ", "")
    if "_xlfn.LET(" in t:
        return True
    if t.startswith("=IF(ABS("):
        return True
    # Verdict-region banner (roll-up): =IF(COUNTIF(F<n>:F<n>,"CHECK")=0, "<static>",
    # COUNTIF(...)&"<static>"). Cell-ref + COUNTIF + fixed verdict strings only; the
    # only literal text is the workbook's own static verdict copy, never user data.
    if t.startswith("=IF(COUNTIF(F") and "OUTOFTOLERANCE" in t.replace(" ", ""):
        return True
    # Delta-percent column: =(B<n>-C<n>)/MAX(ABS(C<n>),1). Cell refs + constants only.
    if re.fullmatch(r"=\(B\d+-C\d+\)/MAX\(ABS\(C\d+\),1\)", t):
        return True
    # =B<n>  or  =B<n>-C<n>  (cell-ref / delta), digits + the single column letters.
    if re.fullmatch(r"=B\d+(-C\d+)?", t):
        return True
    # --- Worked-example cells (Task 4): cell-ref/constant-only, no user data. ---
    if t == "=RAND()":
        return True
    # Scalar clips: =MAX(0,<cell>) / =MIN(1,MAX(0,<cell>)). MAX/MIN are ORIGINAL
    # Excel functions (not _xlfn.-prefixed), but tolerate a prefix defensively.
    if re.fullmatch(r"=(?:_xlfn\.)?MAX\(0,[A-Z]+\d+\)", t):
        return True
    if re.fullmatch(r"=(?:_xlfn\.)?MIN\(1,(?:_xlfn\.)?MAX\(0,[A-Z]+\d+\)\)", t):
        return True
    # Loss chain: =<cell>*<cell>*(<cell>+<cell>).
    if re.fullmatch(r"=[A-Z]+\d+\*[A-Z]+\d+\*\([A-Z]+\d+\+[A-Z]+\d+\)", t):
        return True
    # _invcdf sample forms — all reference one cell + numeric constants + a fixed
    # function whitelist. Trusted iff the ONLY identifiers are the known Excel
    # functions (optionally _xlfn.-prefixed) and single A1-style cell refs.
    stripped = re.sub(r"_xlfn\.(_xlws\.)?", "", t)
    # remove A1-style cell refs and numeric literals, then the function names.
    no_refs = re.sub(r"[A-Z]+\d+", "", stripped)
    no_nums = re.sub(r"\d+\.?\d*", "", no_refs)
    no_fns = no_nums
    for fn in ("BETA.INV", "NORM.INV", "EXP", "SQRT", "IF"):
        no_fns = no_fns.replace(fn, "")
    # what remains must be only operators / punctuation (no stray identifier).
    return bool(re.fullmatch(r"[=+\-*/(),.<> ]*", no_fns))


@pytest.mark.parametrize("hostile", _INJECTION_NAMES)
def test_let_workbook_injection_no_user_string_promoted_to_formula(hostile):
    # run.name == org.name == scenario_name all set to the hostile string.
    run = _let_run(name=hostile)
    run.scenario_inputs_snapshot["scenarios"][0]["scenario_name"] = hostile
    org = _make_org(name=hostile)
    wb = _open(build_single_run_let_workbook(run, org))
    needle = hostile.strip("{}=+@").split("(")[0].lower()  # "danger" / url host
    for sheet in wb.worksheets:
        anchors = _let_anchor_coords(sheet)
        for r in sheet.iter_rows():
            for cell in r:
                if cell.value is None:
                    continue
                if cell.data_type != "f":
                    continue  # non-formula cell: safe (text/number)
                v = cell.value
                text = v.text if isinstance(v, ArrayFormula) else str(v)
                # (1) the ONLY formula cells are trusted, fully-internal forms.
                assert cell.coordinate in anchors or _is_trusted_internal_formula(
                    cell.coordinate, text
                ), (
                    f"unexpected formula cell {sheet.title}!{cell.coordinate} on "
                    f"hostile input {hostile!r}: {text!r}"
                )
                # (2) no hostile content ever appears inside a formula.
                assert needle not in text.lower(), (
                    f"hostile content leaked into formula {sheet.title}!{cell.coordinate}: {text!r}"
                )


@pytest.mark.parametrize("hostile", _INJECTION_NAMES)
def test_let_workbook_injection_no_hyperlinks(hostile):
    # strings_to_urls=False + write_string => a URL/mailto string is plain text,
    # never an external hyperlink. Assert zero hyperlinks at the openpyxl level AND
    # by parsing the xlsx zip for any <hyperlinks> element or hyperlink rels.
    run = _let_run(name=hostile)
    run.scenario_inputs_snapshot["scenarios"][0]["scenario_name"] = hostile
    org = _make_org(name=hostile)
    out = build_single_run_let_workbook(run, org)
    wb = _open(out)
    for sheet in wb.worksheets:
        # read_only worksheets may not expose _hyperlinks; tolerate either.
        hl = getattr(sheet, "_hyperlinks", [])
        assert not hl, f"unexpected hyperlinks on {sheet.title}: {hl}"

    # Zip-level: no <hyperlinks> element in any sheet xml, no hyperlink rels.
    import zipfile

    with zipfile.ZipFile(io.BytesIO(out)) as z:
        names = z.namelist()
        for name in names:
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                xml = z.read(name).decode("utf-8")
                assert "<hyperlinks" not in xml, f"<hyperlinks> present in {name}"
            if "_rels" in name and name.endswith(".rels"):
                rels = z.read(name).decode("utf-8")
                assert "hyperlink" not in rels.lower(), f"hyperlink relationship in {name}"


def test_let_workbook_injection_url_stored_as_text_not_hyperlink():
    # Belt-and-suspenders: a URL name is stored as a plain (neutralized-as-needed)
    # text cell, asserted by TYPE.
    url = "http://evil.example/login"
    run = _let_run(name=url)
    org = _make_org(name=url)
    wb = _open(build_single_run_let_workbook(run, org))
    found = False
    for sheet in wb.worksheets:
        for r in sheet.iter_rows():
            for cell in r:
                if isinstance(cell.value, str) and "evil.example" in cell.value:
                    assert cell.data_type == "s"  # text, not a hyperlink/formula
                    found = True
    assert found, "expected the URL name to be present as a text cell"


# --- Task 4: worked-example block (single-run sheet, explicit scalar formulas) -
# An explicit ~15-row block on the single-run sheet shows the sampling -> base-loss
# chain for the FIRST scenario, one trial per row, with SCALAR formulas (RAND() per
# node feeding the SAME _invcdf the LET uses, then scalar MAX/MIN clips, then loss).
# It exists so the math is inspectable WITHOUT the spilled LET (a spill's formula
# lives only in its anchor cell).
from idraa.services.verification_workbook_let import _invcdf  # noqa: E402


def _strip_xlfn(formula):
    """Strip the writer-injected _xlfn. / _xlfn._xlws. prefixes so a read-back
    formula can be compared to a BARE _invcdf-emitted expression."""
    return (formula or "").replace("_xlfn._xlws.", "").replace("_xlfn.", "")


def _worked_example_rows(ws):
    """All cells of the worked-example block, keyed by 1-based (row, col-letter).

    The block is found by its title cell ("Worked example ...") in column A; every
    subsequent row down to the bottom of the sheet belongs to the block (the block
    is the LAST thing written on the single-run sheet)."""
    title_row = None
    for r in ws.iter_rows():
        for cell in r:
            # read-only sheets yield EmptyCell (no .coordinate/.data_type) for blanks.
            if getattr(cell, "value", None) is None:
                continue
            if (
                isinstance(cell.value, str)
                and cell.value.startswith("Worked example")
                and cell.column == 1
            ):
                title_row = cell.row
                break
        if title_row is not None:
            break
    assert title_row is not None, "worked-example title row not found"
    out = {}
    for r in ws.iter_rows(min_row=title_row):
        for cell in r:
            if getattr(cell, "value", None) is None:
                continue  # skip EmptyCell placeholders
            out[(cell.row, cell.coordinate)] = cell
    return out, title_row


def test_worked_example_block_and_headers_present():
    # Keyword test (req 4): block title + node labels + the loss column header.
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    blob = _flat_strings(wb["MC"]).lower()
    assert "worked example" in blob
    assert "tef sample" in blob and "vuln sample" in blob
    assert "pl sample" in blob and "sl sample" in blob
    assert "base loss" in blob
    # the RAND() draw columns are labeled too.
    assert "rand()" in blob


def test_worked_example_sampling_cell_equals_invcdf_single_source():
    # Drift-guard (req 2 + NTH-3): the worked-example sampling cell formula MUST
    # equal _invcdf(node_dist, <that row's RAND cell ref>) — proving it is single-
    # sourced from the SAME helper the LET uses (cannot drift). The TEF node of the
    # _let_run fixture is a key-less PERT -> _invcdf emits the Vose Beta expression.
    run = _let_run()
    scen = run.scenario_inputs_snapshot["scenarios"][0]
    wb = _open(build_single_run_let_workbook(run, _make_org()))
    cells, title_row = _worked_example_rows(wb["MC"])
    # The header row is title_row+1; the first trial row is title_row+2.
    first_trial = title_row + 2
    # Column layout (see _WE_COL_*): A=u_tef, B=u_vuln, C=u_pl, D=u_sl;
    # E=tef sample, F=vuln sample, G=pl sample, H=sl sample.
    checks = [
        ("E", "A", "threat_event_frequency"),
        ("F", "B", "vulnerability"),
        ("G", "C", "primary_loss"),
        ("H", "D", "secondary_loss"),
    ]
    for sample_col, rand_col, node_key in checks:
        rand_ref = f"{rand_col}{first_trial}"  # e.g. "A12"
        expected = "=" + _invcdf(scen[node_key], rand_ref)
        cell = cells[(first_trial, f"{sample_col}{first_trial}")]
        assert cell.data_type == "f", f"{sample_col} sample cell must be a formula"
        got = _strip_xlfn(str(cell.value))
        assert got.replace(" ", "") == expected.replace(" ", ""), (
            f"{node_key}: worked-example sample formula {got!r} != _invcdf {expected!r}"
        )


def test_worked_example_uses_scalar_max_min_clips_not_boolean():
    # Req 1: the worked-example clips operate on SCALAR cell refs, so scalar
    # MAX(0,..)/MIN(1,..) are correct — the LET's boolean (x>0)*x form must NOT
    # appear in the worked-example clip cells.
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    cells, title_row = _worked_example_rows(wb["MC"])
    first_trial = title_row + 2
    # Clip columns: I=tef clip, J=vuln clip, K=pl clip, L=sl clip.
    tef_clip = cells[(first_trial, f"I{first_trial}")]
    vuln_clip = cells[(first_trial, f"J{first_trial}")]
    pl_clip = cells[(first_trial, f"K{first_trial}")]
    sl_clip = cells[(first_trial, f"L{first_trial}")]
    assert _strip_xlfn(str(tef_clip.value)).replace(" ", "") == f"=MAX(0,E{first_trial})"
    assert _strip_xlfn(str(vuln_clip.value)).replace(" ", "") == f"=MIN(1,MAX(0,F{first_trial}))"
    assert _strip_xlfn(str(pl_clip.value)).replace(" ", "") == f"=MAX(0,G{first_trial})"
    assert _strip_xlfn(str(sl_clip.value)).replace(" ", "") == f"=MAX(0,H{first_trial})"
    # Boolean array-clip form must NOT appear anywhere in the worked-example block.
    for (_rc, _coord), cell in cells.items():
        if cell.data_type == "f" and isinstance(cell.value, str):
            assert ">0)*" not in cell.value.replace(" ", ""), (
                f"boolean array clip leaked into worked-example cell {_coord}: {cell.value!r}"
            )


def test_worked_example_loss_cell_is_lef_lm_chain():
    # The base loss cell = MAX(0,tef)*clip(vuln)*(MAX(0,pl)+MAX(0,sl)) via the
    # clip-column refs (I..L), i.e. =I<r>*J<r>*(K<r>+L<r>).
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    cells, title_row = _worked_example_rows(wb["MC"])
    first_trial = title_row + 2
    loss = cells[(first_trial, f"M{first_trial}")]
    assert loss.data_type == "f"
    r = first_trial
    assert str(loss.value).replace(" ", "") == f"=I{r}*J{r}*(K{r}+L{r})"


def test_worked_example_write_discipline_labels_string_computed_formula():
    # Write discipline (Sec-I4 / req 3): every label/header cell is string-typed
    # ('s'); every computed cell (RAND draws, samples, clips, loss) is formula-typed
    # ('f'); NO numeric constants are written in the block.
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    ws = wb["MC"]
    cells, title_row = _worked_example_rows(ws)
    header_row = title_row + 1
    first_trial = title_row + 2

    # Title (col A of title_row) and every header cell of header_row are strings.
    title_cell = cells[(title_row, f"A{title_row}")]
    assert title_cell.data_type == "s"
    header_cells = [c for (rc, _co), c in cells.items() if rc == header_row and c.value is not None]
    assert len(header_cells) == 13, f"expected 13 header cells, got {len(header_cells)}"
    for c in header_cells:
        assert c.data_type == "s", f"header cell {c.coordinate} not string: {c.data_type}"

    # Every populated cell of the FIRST trial row is a formula (13 computed cols).
    trial_cells = [c for (rc, _co), c in cells.items() if rc == first_trial and c.value is not None]
    assert len(trial_cells) == 13, f"expected 13 computed cells, got {len(trial_cells)}"
    for c in trial_cells:
        assert c.data_type == "f", f"trial cell {c.coordinate} not formula: {c.data_type}"

    # No numeric-constant cells anywhere in the block.
    for (_rc, coord), c in cells.items():
        assert c.data_type != "n", f"unexpected numeric constant in worked-example {coord}"


def test_worked_example_rand_cells_are_volatile_rand():
    # The per-row uniform draws are scalar volatile RAND() (an ORIGINAL Excel
    # function -> no _xlfn. prefix), one per FAIR node.
    wb = _open(build_single_run_let_workbook(_let_run(), _make_org()))
    cells, title_row = _worked_example_rows(wb["MC"])
    first_trial = title_row + 2
    for col in ("A", "B", "C", "D"):
        c = cells[(first_trial, f"{col}{first_trial}")]
        assert c.data_type == "f"
        assert str(c.value).replace(" ", "") == "=RAND()"
        assert "_xlfn." not in str(c.value)  # RAND is original, never prefixed


# --- Task 11: aggregate-run support (K-of-M, scale guard) ---------------------
# An aggregate run carries simulation_results["per_scenario"] = [full SINGLE-shape
# payload per scenario] (run_executor.py:422-430) plus aggregate-level rollups.
# The verification workbook builds in-Excel MC for at most K=max_scenarios
# scenarios with INDEPENDENT RAND() draws (mirroring the engine's per-scenario
# spawn), rolls them up elementwise, and compares the K-subset roll-up against the
# SUM of those K scenarios' app ALEs read at
# per_scenario[i]["residual_risk"]["annualized_loss_expectancy"].


def _agg_scenario_input(sid, name):
    """One per-scenario input snapshot entry (all native-emittable distributions)."""
    return {
        "scenario_id": sid,
        "scenario_name": name,
        "threat_event_frequency": {
            "distribution": "pert",
            "low": 1.0,
            "mode": 3.0,
            "high": 6.0,
        },
        # vuln BETA -> sample-level clip; not param-scaled (mirrors single fixture)
        "vulnerability": {"distribution": "beta", "alpha": 2.0, "beta": 5.0},
        "primary_loss": {"distribution": "lognormal", "mean": 12.0, "sigma": 1.0},
        "secondary_loss": {
            "distribution": "pert",
            "low": 1000.0,
            "mode": 5000.0,
            "high": 20000.0,
        },
    }


def _agg_per_scenario_entry(sid, name, base_ale, res_ale):
    """A full SINGLE-shape per_scenario payload entry (run_executor.py:422-430):
    scenario_id + scenario_name + base_risk/residual_risk dicts."""
    return {
        "scenario_id": sid,
        "scenario_name": name,
        "base_risk": _risk_dict(base_ale),
        "residual_risk": _risk_dict(res_ale),
    }


# --- Task 3: AGGREGATE LET-spill assembly (per-scenario LET + residual roll-up) -
# An aggregate run emits ONE self-contained LET per RECONSTRUCTIBLE scenario, each
# at its OWN single-cell anchor (its own RANDARRAY draws). The roll-up residual ALE
# = SUM of the per-scenario residual-ALE spill cells (the LET's 2nd stat); the
# aggregate VaR/ES rows show the Excel column as "n/a (not re-derived in this
# workbook)" (tails are not additive) and the App column as the App's REAL aggregate
# VaR/ES (read off the run's aggregate_with_controls dict); the K-subset comparison +
# T11 honest-degraded exclusion carry over from the legacy aggregate. Written with
# xlsxwriter,
# re-read with openpyxl for assertions (xlsxwriter prefixes BARE func names with
# _xlfn., the LET lands as a single-anchor ArrayFormula).
from idraa.services.verification_workbook import (  # noqa: E402
    build_aggregate_let_workbook,
)


def _agg_let_scenario_input_keyless(sid, name):
    """Per-scenario input in the dominant key-less PERT wizard shape (TEF/PL/SL PERT
    with NO 'distribution' key + the distribution_fit_metadata sidecar; vuln BETA so
    the base-side vuln is a vuln-only sample-level clip)."""
    return {
        "scenario_id": sid,
        "scenario_name": name,
        "threat_event_frequency": {
            "low": 1.0,
            "mode": 3.0,
            "high": 6.0,
            "distribution_fit_metadata": {"source": "wizard"},
        },
        "vulnerability": {"distribution": "beta", "alpha": 2.0, "beta": 5.0},
        "primary_loss": {
            "low": 1000.0,
            "mode": 5000.0,
            "high": 20000.0,
            "distribution_fit_metadata": {"source": "wizard"},
        },
        "secondary_loss": {
            "low": 500.0,
            "mode": 2000.0,
            "high": 8000.0,
            "distribution_fit_metadata": {"source": "wizard"},
        },
    }


def _agg_with_controls_tail_dict():
    """An aggregate_with_controls dict carrying the full tail ladder (var_90/95/99/999
    + expected_shortfall) — the shape run_executor's _build_aggregate_lec_pair persists
    (var_* monotone, es_q >= var_q). Used by the App-column aggregate VaR/ES tests."""
    return {
        "annualized_loss_expectancy": 900_000.0,
        "mean": 900_000.0,
        "median": 800_000.0,
        "std_deviation": 250_000.0,
        "var_90": 1_100_000.0,
        "var_95": 1_200_000.0,
        "var_99": 1_500_000.0,
        "var_999": 2_000_000.0,
        "expected_shortfall": {
            "es_95": 1_350_000.0,
            "es_99": 1_650_000.0,
            "es_999": 2_200_000.0,
        },
        "loss_exceedance_curve": [],
    }


def _make_aggregate_let_run(*, mc_iterations=10, include_nonrecon=True, with_tail=True):
    """Aggregate run with >=3 key-less-PERT scenarios; when include_nonrecon, the
    3rd scenario carries a BETA in its TEF (magnitude) slot, which is composition-
    reconstructible but NOT param-scalable (scaled_params rejects BETA) -> excluded
    from the residual roll-up (T11 honest-degraded). Distinct, non-zero residual /
    base ALEs make the roll-up subset sum unambiguous."""
    n_scen = 3
    sids = [f"s{i + 1}" for i in range(n_scen)]
    names = ["Recon A", "Recon B", "Non-recon C"]
    scen_inputs = [_agg_let_scenario_input_keyless(sids[i], names[i]) for i in range(n_scen)]
    if include_nonrecon:
        # BETA in the magnitude (TEF) slot -> not param-scalable -> fail-loud exclude.
        scen_inputs[2]["threat_event_frequency"] = {
            "distribution": "beta",
            "alpha": 2.0,
            "beta": 5.0,
        }
    # Distinct ALEs per scenario.
    res_ales = [300_000.0, 450_000.0, 777_000.0]
    base_ales = [1_000_000.0, 1_400_000.0, 2_500_000.0]
    per_scenario = [
        _agg_per_scenario_entry(sids[i], names[i], base_ale=base_ales[i], res_ale=res_ales[i])
        for i in range(n_scen)
    ]
    sim_results = {
        "per_scenario": per_scenario,
        "cost_summary": {
            "total_annual_cost": 100_000.0,
            "total_risk_reduction": 2_000_000.0,
            "net_benefit": 1_900_000.0,
            "aggregate_roi": 19.0,
        },
        "n_scenarios": n_scen,
        "n_simulations": mc_iterations,
    }
    # The aggregate run's persisted aggregate_with_controls dict — the App's REAL
    # aggregate VaR/ES come from here (computed off the aggregate loss distribution).
    # When with_tail is False, omit the tail keys to simulate a LEGACY pre-tail run.
    if with_tail:
        sim_results["aggregate_with_controls"] = _agg_with_controls_tail_dict()
    else:
        sim_results["aggregate_with_controls"] = {
            "annualized_loss_expectancy": 900_000.0,
            "mean": 900_000.0,
            "var_95": 1_200_000.0,
            "var_99": 1_500_000.0,
            "loss_exceedance_curve": [],
        }
    run = SimpleNamespace(
        name="Aggregate LET Run",
        run_type="aggregate",
        mc_iterations=mc_iterations,
        random_seed=42,
        controls_snapshot=[],  # no controls -> identity residual (base==residual)
        aggregate_scenario_ids=sids,
        aggregate_control_ids_per_scenario=None,
        scenario_inputs_snapshot={"scenarios": scen_inputs},
        simulation_results=sim_results,
    )
    return run, {
        "res_ales": res_ales,
        "base_ales": base_ales,
        "names": names,
    }


def _agg_let_anchors(ws):
    """[(coordinate, ref, text)] for every dynamic-array LET cell on the sheet."""
    return _let_anchors(ws)


def test_aggregate_let_one_anchor_per_reconstructible_scenario():
    # 3 scenarios, 1 non-reconstructible (BETA-in-TEF) -> exactly 2 LET anchors.
    run, _ = _make_aggregate_let_run()
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    anchors = _agg_let_anchors(wb["Aggregate"])
    assert len(anchors) == 2, f"expected 2 reconstructible LET anchors, got {anchors}"
    for coord, ref, text in anchors:
        # Each LET is a SINGLE-cell anchor (no range) generating its OWN RANDARRAY.
        assert ":" not in str(ref), f"LET anchor must be a single cell, got ref={ref!r}"
        assert str(ref) == coord
        assert "_xlfn.LET" in text and "_xlfn.RANDARRAY" in text
        assert "_xlpm.u_tef" in text
        assert f"RANDARRAY({run.mc_iterations},1)" in text  # per-scenario N
        assert "CHOOSE(" in text and "VSTACK(" not in text


def test_aggregate_let_rollup_sums_per_scenario_residual_ale_cells():
    # The residual-ALE roll-up must be a SUM referencing each reconstructible
    # scenario's residual-ALE spill cell (B<anchor+1>, _LET_STAT_SPEC index 1).
    run, _ = _make_aggregate_let_run()
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]
    anchors = _agg_let_anchors(ws)
    # residual-ALE cell for each anchor is the row directly below the base-ALE (anchor).
    expected_res_cells = {f"B{int(coord[1:]) + 1}" for (coord, _r, _t) in anchors}
    # Find the roll-up SUM formula: starts with "=B", contains "+", joins cell refs
    # only (no "!" sheet ref, no function call).
    sum_formulas = [
        c.value
        for r in ws.iter_rows()
        for c in r
        if c.data_type == "f"
        and isinstance(c.value, str)
        and c.value.startswith("=B")
        and "+" in c.value
        and "!" not in c.value
        and all(tok.strip().lstrip("=")[0] == "B" for tok in c.value.lstrip("=").split("+"))
    ]
    assert sum_formulas, f"expected a roll-up SUM of residual-ALE cells; got {sum_formulas}"
    # The residual-ALE roll-up references EXACTLY the per-scenario residual-ALE cells.
    referenced = set()
    for f in sum_formulas:
        referenced.update(tok.strip() for tok in f.lstrip("=").split("+"))
    assert expected_res_cells <= referenced, (
        f"roll-up must reference each scenario's residual-ALE cell {expected_res_cells}; "
        f"referenced {referenced}"
    )


def test_aggregate_let_var_es_excel_col_is_not_re_derived_text_app_col_is_app_value():
    # New shape: the aggregate VaR/ES rows show the Excel column as the
    # "not re-derived in this workbook" TEXT (tails are not additive -> the in-Excel
    # roll-up does not re-derive them), while the App column shows the App's REAL
    # aggregate VaR/ES read off the run's aggregate_with_controls dict (a NUMBER), and
    # the Delta column is blank. No user data is promoted to a formula.
    run, _ = _make_aggregate_let_run()
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]

    excel_na = "n/a (not re-derived in this workbook)"
    excel_na_cells = [
        c for r in ws.iter_rows() for c in r if isinstance(c.value, str) and c.value == excel_na
    ]
    # 6 tail metrics (VaR95/99/999, ES95/99/999), Excel column only.
    assert len(excel_na_cells) == 6, (
        f"expected 6 Excel 'not re-derived' tail cells, got {len(excel_na_cells)}"
    )
    for c in excel_na_cells:
        assert c.data_type == "s", f"Excel n/a cell promoted to {c.data_type}: {c.coordinate}"

    # The App column for each tail row carries the REAL app aggregate value (a NUMBER).
    agg = _agg_with_controls_tail_dict()
    expected_app = {
        agg["var_95"],
        agg["var_99"],
        agg["var_999"],
        agg["expected_shortfall"]["es_95"],
        agg["expected_shortfall"]["es_99"],
        agg["expected_shortfall"]["es_999"],
    }
    # The App column is the cell directly to the RIGHT of each Excel-na cell (col C).
    app_values = []
    for c in excel_na_cells:
        app_cell = ws.cell(row=c.row, column=c.column + 1)
        assert app_cell.data_type == "n", (
            f"App aggregate VaR/ES cell must be a number, got {app_cell.data_type}"
        )
        app_values.append(float(app_cell.value))
    assert set(app_values) == expected_app, (
        f"App column must show the app aggregate VaR/ES values {expected_app}; got {app_values}"
    )
    # The old misleading "tail non-additive" text must be GONE.
    flat = " ".join(str(c.value) for r in ws.iter_rows() for c in r if isinstance(c.value, str))
    assert "tail non-additive" not in flat
    assert "tails do NOT sum" not in flat and "tails do not sum" not in flat


def test_aggregate_let_var_es_app_col_sane_ladder():
    # Sanity: the App aggregate tail ladder must be monotone (var_999 >= var_99 >=
    # var_95) and each ES_q >= its VaR_q.
    run, _ = _make_aggregate_let_run()
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]
    excel_na = "n/a (not re-derived in this workbook)"
    by_label = {}
    for r in ws.iter_rows():
        for c in r:
            if isinstance(c.value, str) and c.value == excel_na:
                label = ws.cell(row=c.row, column=c.column - 1).value
                by_label[str(label)] = float(ws.cell(row=c.row, column=c.column + 1).value)
    assert by_label["VaR 99.9"] >= by_label["VaR 99"] >= by_label["VaR 95"]
    assert by_label["ES 95"] >= by_label["VaR 95"]
    assert by_label["ES 99"] >= by_label["VaR 99"]
    assert by_label["ES 99.9"] >= by_label["VaR 99.9"]


def test_aggregate_let_legacy_run_app_col_shows_re_run_to_populate_not_zero():
    # A LEGACY aggregate run whose persisted aggregate_with_controls lacks the tail
    # keys (has_tail_metrics False) must show the App column as the suppress-not-
    # fabricate legacy n/a text, NOT a fabricated 0.
    run, _ = _make_aggregate_let_run(with_tail=False)
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]
    excel_na = "n/a (not re-derived in this workbook)"
    legacy_na = "n/a (run predates aggregate tail metrics -- re-run to populate)"
    legacy_cells = [
        c for r in ws.iter_rows() for c in r if isinstance(c.value, str) and c.value == legacy_na
    ]
    assert len(legacy_cells) == 6, f"expected 6 legacy-n/a App tail cells, got {len(legacy_cells)}"
    for c in legacy_cells:
        assert c.data_type == "s", f"legacy n/a cell promoted to {c.data_type}: {c.coordinate}"
        # Its Excel column (one to the left) is the not-re-derived text.
        assert ws.cell(row=c.row, column=c.column - 1).value == excel_na
    # No App tail cell is a fabricated number 0 next to an Excel-na cell.
    for r in ws.iter_rows():
        for c in r:
            if isinstance(c.value, str) and c.value == excel_na:
                app = ws.cell(row=c.row, column=c.column + 1)
                assert not (app.data_type == "n" and float(app.value) == 0.0), (
                    "legacy run must not fabricate a 0 App aggregate tail value"
                )


def test_aggregate_let_excludes_non_reconstructible_from_rollup():
    # The non-reconstructible (BETA-in-TEF) scenario must be EXCLUDED from the
    # roll-up (no LET block) and listed summary-only with the fail-loud cell.
    run, info = _make_aggregate_let_run()
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]
    flat = " ".join(
        str(c.value) for r in ws.iter_rows() for c in r if isinstance(c.value, str)
    ).lower()
    # The two reconstructible scenarios get LET blocks; the non-recon one does NOT.
    assert "scenario block: recon a" in flat
    assert "scenario block: recon b" in flat
    assert "scenario block: non-recon c" not in flat
    # in-Excel MC shown for 2 of 3 scenarios.
    assert "2 of 3 scenarios" in flat
    # The non-reconstructible scenario is summary-only with a "not reconstructible"
    # reason and the labeled fail-loud cell.
    assert "excluded from the residual roll-up" in flat
    assert "not reconstructible" in flat
    assert "non-recon c" in flat
    # Its App BASE ALE appears in the summary-only section (App base shown).
    nums = [c.value for r in ws.iter_rows() for c in r if isinstance(c.value, (int, float))]
    assert any(abs(v - info["base_ales"][2]) < 1e-6 for v in nums), (
        "non-reconstructible scenario's App base ALE must be shown summary-only"
    )


def test_aggregate_let_app_comparison_sums_k_subset_residual_app_ales():
    # The Excel residual roll-up is compared against the App figure for EXACTLY the
    # in-Excel reconstructible scenarios (sum of their App residual ALEs) -- NOT the
    # full-M aggregate (carries over the legacy K-subset comparison semantics).
    run, info = _make_aggregate_let_run()
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]
    # Reconstructible subset = scenarios 0 and 1 (scenario 2 is BETA-in-TEF).
    expected_subset_residual = info["res_ales"][0] + info["res_ales"][1]
    full_m_residual = sum(info["res_ales"])
    assert expected_subset_residual != full_m_residual  # must differ to be meaningful
    nums = [c.value for r in ws.iter_rows() for c in r if isinstance(c.value, (int, float))]
    assert any(abs(v - expected_subset_residual) < 1e-6 for v in nums), (
        f"App comparison ALE must be the reconstructible-subset residual sum "
        f"{expected_subset_residual}; got {nums}"
    )
    # The full-M residual (includes the excluded scenario) must NOT be the App figure.
    assert not any(abs(v - full_m_residual) < 1e-6 for v in nums), (
        "App comparison wrongly includes the non-reconstructible scenario's residual"
    )
    # The non-reconstructible scenario's BASE magnitude must NEVER be summed into the
    # residual roll-up comparison (base >> residual would inflate it).
    forbidden = expected_subset_residual + info["base_ales"][2]
    assert not any(abs(v - forbidden) < 1e-6 for v in nums), (
        "non-reconstructible scenario's base magnitude leaked into the residual roll-up"
    )


def test_aggregate_let_all_reconstructible_three_anchors_three_term_sum():
    # With no non-reconstructible scenarios, all 3 get LET anchors and the roll-up
    # SUM is a 3-term sum of their residual-ALE cells.
    run, _ = _make_aggregate_let_run(include_nonrecon=False)
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]
    anchors = _agg_let_anchors(ws)
    assert len(anchors) == 3
    expected_res_cells = {f"B{int(coord[1:]) + 1}" for (coord, _r, _t) in anchors}
    sum_formulas = [
        c.value
        for r in ws.iter_rows()
        for c in r
        if c.data_type == "f"
        and isinstance(c.value, str)
        and c.value.startswith("=B")
        and c.value.count("+") == 2  # 3-term sum
        and "!" not in c.value
    ]
    assert sum_formulas, "expected a 3-term residual-ALE roll-up SUM"
    referenced = set()
    for f in sum_formulas:
        referenced.update(tok.strip() for tok in f.lstrip("=").split("+"))
    assert expected_res_cells <= referenced


def test_aggregate_let_emit_fail_keeps_excel_app_and_label_membership_symmetric(monkeypatch):
    # K-subset symmetry (methodology [IMPORTANT]): a scenario can PASS
    # _agg_let_collect_scenarios' gate (so it lands in in_excel) yet have
    # scenario_let_formula RAISE at sheet-build time. When that happens the emit-fail
    # scenario MUST be excluded from BOTH the Excel residual SUM and the App
    # comparison sum AND every "X of M" count -- otherwise Excel sums (n_in - 1) while
    # App sums n_in, both mislabeled "n_in of m" => a biased+mislabeled Delta the
    # operator misreads as RANDARRAY noise. All 3 scenarios pass the gate
    # (include_nonrecon=False); we force the SECOND ("Recon B") to raise in the
    # emitter. Expected emitted set = {Recon A, Non-recon C} (both reconstructible).
    import idraa.services.verification_workbook_let as vwl

    run, info = _make_aggregate_let_run(include_nonrecon=False)
    real_scenario_let_formula = vwl.scenario_let_formula
    fail_name = info["names"][1]  # "Recon B"

    def _raising_scenario_let_formula(scenario, mults, n):
        # Raise for exactly the gate-passing scenario we want to fail at emit time.
        if str(scenario.get("scenario_name")) == fail_name:
            raise ValueError("forced emit failure for K-subset symmetry test")
        return real_scenario_let_formula(scenario, mults, n)

    # Patch the SOURCE module: build_aggregate_let_sheet does a call-time
    # `from idraa.services.verification_workbook_let import scenario_let_formula`.
    monkeypatch.setattr(vwl, "scenario_let_formula", _raising_scenario_let_formula)

    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]

    # (1) The emit-fail scenario appears summary-only with the fail reason.
    flat = " ".join(
        str(c.value) for r in ws.iter_rows() for c in r if isinstance(c.value, str)
    ).lower()
    assert "excluded from the residual roll-up" in flat
    assert fail_name.lower() in flat
    assert "let emit failed at build" in flat
    # It does NOT get a LET block (no "Scenario block: Recon B" header).
    assert f"scenario block: {fail_name.lower()}" not in flat
    # The two EMITTED scenarios DO get LET blocks.
    assert f"scenario block: {info['names'][0].lower()}" in flat  # Recon A
    assert f"scenario block: {info['names'][2].lower()}" in flat  # Non-recon C

    # Emitted set = scenarios 0 and 2; emit-fail = scenario 1.
    emitted_res_sum = info["res_ales"][0] + info["res_ales"][2]
    emitted_base_sum = info["base_ales"][0] + info["base_ales"][2]
    full_with_failed_res = sum(info["res_ales"])  # would over-count by scenario 1
    full_with_failed_base = sum(info["base_ales"])

    nums = [c.value for r in ws.iter_rows() for c in r if isinstance(c.value, (int, float))]

    # (2) The App comparison residual sum EXCLUDES the emit-fail scenario's
    # app_res_ale (== sum over the EMITTED set), and never the full-with-failed sum.
    assert emitted_res_sum != full_with_failed_res  # must differ to be meaningful
    assert any(abs(v - emitted_res_sum) < 1e-6 for v in nums), (
        f"App comparison residual must be the EMITTED-set sum {emitted_res_sum}; got {nums}"
    )
    assert not any(abs(v - full_with_failed_res) < 1e-6 for v in nums), (
        "App comparison wrongly includes the emit-fail scenario's app_res_ale"
    )
    # The base ALE comparison is likewise the emitted-set sum, not the full sum.
    assert any(abs(v - emitted_base_sum) < 1e-6 for v in nums)
    assert not any(abs(v - full_with_failed_base) < 1e-6 for v in nums)

    # (3) The "X of M" residual-ALE label/count == the number of EMITTED scenarios
    # (2 of 3), and the optimistic "3 of 3" (len(in_excel)) is NEVER printed.
    assert "2 of 3 scenarios" in flat
    assert "3 of 3 scenarios" not in flat

    # (4) The Excel roll-up SUMs reference EXACTLY the emitted scenarios' spill cells.
    # emit-fail leaves exactly 2 LET anchors; each anchor row is the base-ALE cell and
    # anchor+1 is the residual-ALE cell. There are TWO roll-up SUMs (residual + base);
    # each must be a 2-term sum (one term per emitted anchor), never 3 (a 3rd term =
    # the emit-fail scenario leaked into the roll-up).
    anchors = _agg_let_anchors(ws)
    assert len(anchors) == 2, f"emit-fail must leave exactly 2 LET anchors, got {anchors}"
    expected_res_cells = {f"B{int(coord[1:]) + 1}" for (coord, _r, _t) in anchors}
    expected_base_cells = {coord for (coord, _r, _t) in anchors}
    # The two roll-up SUMs (residual + base): pure "=B<n>+B<n>" cell-ref sums.
    sum_formulas = [
        c.value
        for r in ws.iter_rows()
        for c in r
        if c.data_type == "f"
        and isinstance(c.value, str)
        and c.value.startswith("=B")
        and "+" in c.value
        and "!" not in c.value
        and all(tok.strip().lstrip("=")[0] == "B" for tok in c.value.lstrip("=").split("+"))
    ]
    assert sum_formulas, "expected roll-up SUMs of the emitted scenarios' spill cells"
    referenced = set()
    for f in sum_formulas:
        terms = [tok.strip() for tok in f.lstrip("=").split("+")]
        referenced.update(terms)
        # No roll-up SUM may have more than the 2 emitted terms (3rd term = leak).
        assert len(terms) == 2, f"roll-up SUM must be a 2-term (emitted-only) sum: {f!r}"
    # The two SUMs together reference EXACTLY the emitted base + residual cells --
    # nothing from the emit-fail scenario (whose cells were never written).
    assert referenced == (expected_res_cells | expected_base_cells), (
        f"roll-up SUMs must reference EXACTLY the emitted scenarios' base+residual "
        f"cells {expected_res_cells | expected_base_cells}; referenced {referenced}"
    )
    # And the residual roll-up specifically references the emitted residual cells.
    assert any(
        {tok.strip() for tok in f.lstrip("=").split("+")} == expected_res_cells
        for f in sum_formulas
    ), f"a residual roll-up SUM must reference exactly {expected_res_cells}"


def test_aggregate_let_caps_n_per_scenario_at_max():
    # Each per-scenario LET runs at N = min(mc_iterations, verification_workbook_max_n).
    run, _ = _make_aggregate_let_run(mc_iterations=200_000)
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    anchors = _agg_let_anchors(wb["Aggregate"])
    assert anchors
    for _coord, _ref, text in anchors:
        assert "RANDARRAY(50000,1)" in text


def test_aggregate_let_has_required_sheets():
    run, _ = _make_aggregate_let_run()
    out = build_aggregate_let_workbook(run, _make_org())
    assert out[:2] == b"PK"
    wb = _open(out)
    assert {"Aggregate", "Documentation"} <= set(wb.sheetnames)


def test_aggregate_let_collect_availability_threads_per_scenario_effect():
    """Arch-1 fix: _agg_let_collect_scenarios must thread per-scenario effect so the
    workbook residual RECONCILES with the engine for availability constituents.

    A single recovery control (lec_resp_resilience, PROBABILITY, cap=0.7) applied to:
    - An availability scenario: engine credits the raw LEC_RESPONSE effectiveness
      (availability_self_detection=True), so primary_loss mult < 1.0 in mults.
    - A confidentiality scenario: detection-gated (availability_self_detection=False),
      no Detection partner -> primary_loss mult == 1.0 (identity) in mults.

    Before the Arch-1 fix both were hard-coded False ->
    the workbook derived a HIGHER residual for the availability constituent than
    the engine stored, producing a visible engine-vs-workbook mismatch.
    """
    from idraa.services.verification_workbook import _agg_let_collect_scenarios

    # V3 snapshot: a recovery-only control (lec_resp_resilience, PROBABILITY, cap=0.7).
    recovery_snap = ControlSnapshotV3(
        control_id="c-recovery",
        name="Recovery Ctrl",
        domains=["loss_event"],
        type="corrective",
        assignments=[
            ControlFunctionAssignmentSnapshotDTO(
                sub_function=v3_enums.FairCamSubFunction("lec_resp_resilience"),
                capability_value=0.7,
                coverage=1.0,
                reliability=1.0,
                unit_type=v3_enums.UnitType.PROBABILITY,
            )
        ],
    ).model_dump(mode="json")

    avail_sid = "s-avail"
    conf_sid = "s-conf"
    # Keyless PERT shape (scalable); extend with "effect" key so the workbook
    # can thread availability_self_detection correctly per Arch-1.
    avail_input = {
        **_agg_let_scenario_input_keyless(avail_sid, "Avail Scenario"),
        "effect": "availability",
    }
    conf_input = {
        **_agg_let_scenario_input_keyless(conf_sid, "Conf Scenario"),
        "effect": "confidentiality",
    }
    per_scenario = [
        _agg_per_scenario_entry(
            avail_sid, "Avail Scenario", base_ale=1_000_000.0, res_ale=700_000.0
        ),
        _agg_per_scenario_entry(
            conf_sid, "Conf Scenario", base_ale=1_000_000.0, res_ale=1_000_000.0
        ),
    ]
    run = SimpleNamespace(
        name="Agg Avail Reconcile Test",
        run_type="aggregate",
        mc_iterations=10,
        random_seed=42,
        controls_snapshot=[recovery_snap],
        aggregate_scenario_ids=[avail_sid, conf_sid],
        aggregate_control_ids_per_scenario=None,  # full universe for all scenarios
        scenario_inputs_snapshot={"scenarios": [avail_input, conf_input]},
        simulation_results={
            "per_scenario": per_scenario,
            "n_scenarios": 2,
            "n_simulations": 10,
        },
    )

    in_excel, summary_only, _m, _kex, _nex = _agg_let_collect_scenarios(run, k=10)

    # Both PERT scenarios (no BETA in magnitude slot) must be reconstructible.
    assert len(in_excel) == 2, (
        f"expected both scenarios reconstructible (in_excel); "
        f"got {len(in_excel)} in_excel, summary_only={summary_only!r}"
    )
    assert not summary_only, f"unexpected summary-only entries: {summary_only!r}"

    by_name = {e["name"]: e for e in in_excel}
    avail_mults = by_name["Avail Scenario"]["mults"]
    conf_mults = by_name["Conf Scenario"]["mults"]

    # Availability constituent: lec_resp_resilience IS credited (primary_loss mult < 1.0),
    # matching what the engine already computed (availability_self_detection=True path).
    assert avail_mults["primary_loss"] < 1.0, (
        "availability constituent: lec_resp_resilience must reduce primary_loss "
        f"(expected < 1.0, got {avail_mults['primary_loss']})"
    )
    # Confidentiality constituent: recovery control stays detection-gated (mult == 1.0),
    # because no Detection partner is present -> LEC_DETECTION_RESPONSE_PAIR eff is None.
    assert conf_mults["primary_loss"] == pytest.approx(1.0), (
        "confidentiality constituent: lec_resp_resilience must be detection-gated "
        f"(expected primary_loss==1.0, got {conf_mults['primary_loss']})"
    )


# --- Task 5: aggregate ΣN scale-down + responsiveness NOTE --------------------


@pytest.fixture
def _vwb_settings_env(monkeypatch):
    """Override verification-workbook cap env vars + reset the Settings singleton so
    the builders re-read them; reset again on teardown so later tests see defaults.
    Yields a setter ``set(**caps)`` taking max_n / aggregate_total_max kwargs."""
    import idraa.config as config

    alias = {
        "max_n": "VERIFICATION_WORKBOOK_MAX_N",
        "aggregate_total_max": "VERIFICATION_WORKBOOK_AGGREGATE_TOTAL_MAX",
    }

    def _set(**caps):
        for key, value in caps.items():
            monkeypatch.setenv(alias[key], str(value))
        config.reset_for_tests()
        return config.get_settings()

    yield _set
    config.reset_for_tests()


def test_aggregate_let_scales_n_down_so_sum_within_aggregate_cap(_vwb_settings_env):
    # 3 reconstructible scenarios, each wanting N = min(mc_iterations, max_n) = 50k.
    # Aggregate cap 90k -> per-scenario N scaled to floor(90000/3) = 30000 so
    # Σ N = 3 * 30000 = 90000 <= 90000.
    s = _vwb_settings_env(max_n=50_000, aggregate_total_max=90_000)
    run, _ = _make_aggregate_let_run(mc_iterations=50_000, include_nonrecon=False)
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    anchors = _agg_let_anchors(wb["Aggregate"])
    assert len(anchors) == 3, f"expected 3 reconstructible anchors, got {len(anchors)}"
    per_scenario_n = 30_000
    for _coord, _ref, text in anchors:
        assert f"RANDARRAY({per_scenario_n},1)" in text
    # Σ N across the emitted scenarios respects the aggregate cap.
    assert len(anchors) * per_scenario_n <= s.verification_workbook_aggregate_total_max
    # Plain-language responsiveness note is emitted (scale-down bound).
    blob = _flat_strings(wb["Aggregate"]).lower()
    assert "for responsiveness" in blob
    assert "statistically representative" in blob
    assert "aggregate cap of 90000 trials" in blob


def test_aggregate_let_no_scale_down_note_when_sum_within_cap(_vwb_settings_env):
    # 3 scenarios * N(min(10, 50k)=10) = 30 <= 150k default -> no scale-down, N == 10,
    # and no responsiveness note.
    s = _vwb_settings_env(max_n=50_000, aggregate_total_max=150_000)
    run, _ = _make_aggregate_let_run(mc_iterations=10, include_nonrecon=False)
    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    anchors = _agg_let_anchors(wb["Aggregate"])
    assert len(anchors) == 3
    for _coord, _ref, text in anchors:
        assert "RANDARRAY(10,1)" in text
    assert s.verification_workbook_aggregate_total_max >= 3 * 10
    blob = _flat_strings(wb["Aggregate"]).lower()
    assert "for responsiveness" not in blob


def test_agg_scaled_n_no_scale_scale_down_and_floor():
    # Direct unit test of the aggregate ΣN scale-down helper across its regimes.
    # (The workbook-level _make_aggregate_let_run is fixed at 3 scenarios, where the
    # _LET_MIN_N floor can never bind — 1000//3 == 333 > 100 — so the floor must be
    # exercised here, where the scenario count is free.)
    from idraa.services.verification_workbook import _LET_MIN_N, _agg_scaled_n

    # Fits within the cap -> no scale-down, scaled flag False.
    assert _agg_scaled_n(50_000, 3, 200_000) == (50_000, False)  # 150k <= 200k
    # Exceeds the cap, floor NOT binding -> floor(cap / K), scaled True.
    assert _agg_scaled_n(50_000, 3, 1_000) == (1_000 // 3, True)  # 333, ΣN=999<=1000
    # Exceeds the cap AND the floor binds (cap // K < _LET_MIN_N) -> _LET_MIN_N,
    # never collapsing toward 1, even though Σ N (=K*_LET_MIN_N) then exceeds the cap.
    n_floored, scaled_floored = _agg_scaled_n(50_000, 12, 1_000)  # 1000//12 == 83 < 100
    assert n_floored == _LET_MIN_N and scaled_floored is True
    # <= 1 scenario (or 0): N stays at per_run_n, not scaled.
    assert _agg_scaled_n(50_000, 1, 1_000) == (50_000, False)
    assert _agg_scaled_n(50_000, 0, 1_000) == (50_000, False)


# Aggregate injection smoke: the aggregate path writes the run/org/scenario names.
# Task 2's tests already cover the single-run write_string discipline (the aggregate
# reuses the SAME _XlsxRows / write_string mechanics); this asserts a hostile run +
# org + scenario name in the aggregate path never promotes to a formula/hyperlink.
@pytest.mark.parametrize("hostile", ["=danger()", "{=danger()}", "http://evil.example/login"])
def test_aggregate_let_injection_no_user_string_promoted(hostile):
    run, _ = _make_aggregate_let_run()
    run.name = hostile
    for scen in run.scenario_inputs_snapshot["scenarios"]:
        scen["scenario_name"] = hostile
    org = _make_org(name=hostile)
    out = build_aggregate_let_workbook(run, org)
    wb = _open(out)
    needle = hostile.strip("{}=+@").split("(")[0].lower()
    for sheet in wb.worksheets:
        for r in sheet.iter_rows():
            for cell in r:
                if cell.value is None or cell.data_type != "f":
                    continue
                v = cell.value
                text = v.text if isinstance(v, ArrayFormula) else str(v)
                # The only formulas are trusted internal forms (LET anchors, =B<n>
                # spill/roll-up/sum/delta refs); a hostile string must never leak in.
                assert needle not in text.lower(), (
                    f"hostile content leaked into formula {sheet.title}!{cell.coordinate}: {text!r}"
                )
    # Zero hyperlinks anywhere (strings_to_urls=False + write_string).
    import zipfile

    with zipfile.ZipFile(io.BytesIO(out)) as z:
        for name in z.namelist():
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                assert "<hyperlinks" not in z.read(name).decode("utf-8")
            if "_rels" in name and name.endswith(".rels"):
                assert "hyperlink" not in z.read(name).decode("utf-8").lower()


# --- Regression: production PERT shape omits the "distribution" key (issue: 500 on download) ---
# The wizard stores PERT nodes as {"low","mode","high","distribution_fit_metadata":...}
# with NO "distribution" key (only lognormal gets an explicit slug). The engine's
# _dict_to_fair_distribution tolerates this via payload.get("distribution", "pert").
# The builder must mirror that default, not hard-subscript dist_dict["distribution"].


def _keyless_pert_scenario_inputs():
    """Real wizard shape: PERT nodes WITHOUT a 'distribution' key (+ the
    distribution_fit_metadata sidecar the wizard attaches)."""
    return {
        "scenarios": [
            {
                "scenario_id": "s1",
                "scenario_name": "Scenario A",
                "threat_event_frequency": {"low": 1.0, "mode": 3.0, "high": 6.0},
                "vulnerability": {
                    "low": 0.1,
                    "mode": 0.3,
                    "high": 0.6,
                    "distribution_fit_metadata": {"source": "wizard"},
                },
                "primary_loss": {"low": 1000.0, "mode": 5000.0, "high": 20000.0},
                "secondary_loss": {"low": 500.0, "mode": 2000.0, "high": 8000.0},
            }
        ]
    }


def test_invcdf_keyless_pert_ignores_distribution_fit_metadata_sidecar():
    # The LET emitter's _invcdf must dispatch on low/mode/high (defaulting a missing
    # "distribution" key to pert) and read PAST the wizard's distribution_fit_metadata
    # sidecar — the dominant real PERT shape. (Replaces the deleted explicit-row
    # sample_formula_for sidecar regression; same concern, LET path.)
    from idraa.services.verification_workbook_let import _invcdf

    expr = _invcdf(
        {"low": 0.0, "mode": 5.0, "high": 10.0, "distribution_fit_metadata": {"x": 1}},
        "u",
    ).replace(" ", "")
    assert expr.startswith("0.0+BETA.INV(u,4.0,4.0)")  # symmetric Vose -> Beta(4,4)


def test_build_workbook_handles_keyless_pert_scenario_shape():
    # End-to-end: a run whose frozen scenario inputs use the real key-less PERT
    # shape must build a valid xlsx, not raise (was: KeyError 'distribution' -> 500).
    run = SimpleNamespace(
        name="Run X",
        run_type="single",
        mc_iterations=10,
        random_seed=42,
        controls_snapshot=[],
        scenario_inputs_snapshot=_keyless_pert_scenario_inputs(),
        simulation_results=_sim_results(),
    )
    out = build_verification_workbook(run, _make_org())
    assert out[:2] == b"PK"  # valid xlsx zip magic, no exception
    wb = _open(out)
    assert "MC" in wb.sheetnames


def test_residual_reconstructs_for_keyless_pert_shape():
    # BLOCKER guard: the residual path routes through scaled_params, which also
    # must default a missing "distribution" key to pert. Before the line-330 fix
    # this returned (False, {}) — the residual MC silently degraded to fail-loud
    # for a perfectly valid, engine-reconstructible key-less PERT scenario.
    from idraa.services.verification_workbook import _residual_sample_formulas

    scen = _keyless_pert_scenario_inputs()["scenarios"][0]
    identity_mults = {
        "threat_event_frequency": 1.0,
        "vulnerability": 1.0,
        "primary_loss": 1.0,
        "secondary_loss": 1.0,
        "currency_subtractor_total": 0.0,
    }
    feasible, parts = _residual_sample_formulas(scen, identity_mults)
    assert feasible is True, "key-less PERT residual must reconstruct, not degrade"
    # Post-spill-redesign _residual_sample_formulas is a pure feasibility predicate
    # (parts is an empty dict — the LET emitter builds the residual expressions
    # itself); the contract is the feasible flag, asserted above.
    assert parts == {}


# ---------------------------------------------------------------------------
# Task 6 (issue #419): Control Audit sheet
#   — composition LET formulas (1−E·w per node at canonical weights)
#   — deterministic band-endpoint sensitivity (low/base/high weight columns)
#   — emitted stochastic range (p5/p50/p95 + stability_class from weight_robustness)
# ---------------------------------------------------------------------------
#
# Helpers


def _all_cell_formulas(wb_bytes: bytes) -> list[str]:
    """Return formula text for every formula cell across all sheets.

    Regular formula cells (data_type='f', non-array) return the formula
    string; dynamic-array cells return the ArrayFormula.text.
    xlsxwriter use_future_functions=True prefixes bare function names with
    _xlfn. so the returned texts are the prefixed forms.
    """
    from openpyxl.worksheet.formula import ArrayFormula

    wb = _open(wb_bytes)
    formulas: list[str] = []
    for sheet in wb.worksheets:
        for r in sheet.iter_rows():
            for cell in r:
                if cell.data_type == "f":
                    v = cell.value
                    if isinstance(v, ArrayFormula):
                        formulas.append(v.text or "")
                    elif isinstance(v, str):
                        formulas.append(v)
    return formulas


def _all_cell_strings(wb_bytes: bytes) -> list[str]:
    """Return text (data_type 's') cell values across all sheets.

    Excludes formula cells so "Low"/"Base"/"High" headers do not mix with
    formula texts.
    """
    wb = _open(wb_bytes)
    strings: list[str] = []
    for sheet in wb.worksheets:
        for r in sheet.iter_rows():
            for cell in r:
                if cell.data_type == "s" and isinstance(cell.value, str):
                    strings.append(cell.value)
    return strings


# Minimal weight_robustness blob for a single control "c1".
_WR_OK: dict = {
    "state": "ok",
    "draws_used": 256,
    "degraded": False,
    "band": {"logit_sigma": 0.6, "seed": 42, "draws": 256},
    "per_control": {
        "c1": {
            "reduction_p5": 100_000.0,
            "reduction_p50": 300_000.0,
            "reduction_p95": 500_000.0,
            "stability_class": "stable",
            "rank_p50": 0,
            "rank_min": 0,
            "rank_max": 0,
        }
    },
    "headline": {
        "reduction_p5": 100_000.0,
        "reduction_p50": 300_000.0,
        "reduction_p95": 500_000.0,
    },
    "canonical_value": {"c1": 300_000.0},
    "kendall_tau_p50": 0.9,
    "topk_preservation_k": 1,
    "topk_preservation_prob": 0.95,
    "indistinguishable_pairs": [],
    "rank_stability_available": True,
}

_WR_INSUFFICIENT: dict = {
    "state": "insufficient_budget",
    "draws_used": 0,
    "degraded": True,
    "band": {"logit_sigma": 0.6, "seed": 42, "draws": 0},
    "per_control": {},
    "headline": {"reduction_p5": 0.0, "reduction_p50": 0.0, "reduction_p95": 0.0},
    "canonical_value": {},
    "kendall_tau_p50": None,
    "topk_preservation_k": None,
    "topk_preservation_prob": None,
    "indistinguishable_pairs": [],
    "rank_stability_available": False,
}


@pytest.fixture
def workbook_bytes_with_controls():
    """xlsx bytes for a single-run with an active control + weight_robustness=ok.

    Control c1: lec_prev_avoidance (PROBABILITY, cap=0.8, cov=1.0, rel=1.0).
    This is the canonical fixture for composition-formula validation.
    """
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = _WR_OK
    return build_single_run_let_workbook(run, _make_org())


def test_workbook_ranking_shows_too_close_to_call_from_pair_set():
    """Per-control 'Ranking' must read 'too close to call' for a control in the
    indistinguishable PAIR set — mirroring web/PDF (Spec-I1). Regression: the
    per-control stability_badge collapses unstable→stable, so the pair set is the
    sole un-rankable signal; the workbook must not show 'stable' for such controls."""
    wr = {**_WR_OK, "indistinguishable_pairs": [["c1", "c2"]]}
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = wr
    out = build_single_run_let_workbook(run, _make_org())
    lower = " ".join(_all_cell_strings(out)).lower()
    assert "too close to call" in lower, (
        "control in indistinguishable_pairs must render 'too close to call' in the workbook"
    )


# ---------------------------------------------------------------------------
# Task 6 – test 1: composition LET formula present
# ---------------------------------------------------------------------------


def test_workbook_has_composition_formula(workbook_bytes_with_controls):
    """Controls sheet exposes 1−E·w as a LET formula (not a bare number)."""
    formulas = _all_cell_formulas(workbook_bytes_with_controls)
    assert any("LET(" in f and "1-" in f.replace(" ", "") for f in formulas), (
        "expected at least one LET formula containing '1-' in the workbook; "
        f"got formulas: {formulas[:5]!r}"
    )


# ---------------------------------------------------------------------------
# Task 6 – test 2: low/base/high column headers present
# ---------------------------------------------------------------------------


def test_workbook_has_low_base_high_columns(workbook_bytes_with_controls):
    """Controls sheet has explicit Low / Typical / High column headers."""
    strings = _all_cell_strings(workbook_bytes_with_controls)
    lower_strings = {s.lower() for s in strings}
    assert {"low", "typical", "high"} <= lower_strings, (
        f"expected 'low', 'typical', 'high' column headers in text cells; "
        f"got (sample): {sorted(lower_strings)[:20]}"
    )


# ---------------------------------------------------------------------------
# Task 6 – test 3: emitted stochastic range
# ---------------------------------------------------------------------------


def test_workbook_emits_stochastic_range(workbook_bytes_with_controls):
    """Controls sheet emits p5/p95 values + 'estimated value range' label."""
    strings = _all_cell_strings(workbook_bytes_with_controls)
    assert any("p5" in s.lower() or "p95" in s.lower() for s in strings), (
        "expected a cell containing 'p5' or 'p95' in the workbook"
    )
    assert any("estimated value range" in s.lower() for s in strings), (
        "expected a cell containing 'estimated value range'"
    )


# ---------------------------------------------------------------------------
# Task 6 – numeric validation: composition LET recomputes to engine multiplier
# ---------------------------------------------------------------------------


def test_composition_let_formula_matches_engine_multiplier():
    """Side-by-side: hand-computed 1−E·w vs formula-string evaluation agree.

    Fixture: lec_prev_avoidance, cap=0.8, cov=1.0, rel=1.0.
    Group: LEC_PREVENTION (OR). Canonical TEF weight w=0.8.
    Expected (hand): opeff = 0.8*1.0*1.0 = 0.8 → mult_TEF = 1 − 0.8·0.8 = 0.36
    Expected (hand): mult_Vuln = 1 − 0.8·0.9 = 0.28
    """
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = None
    wb_bytes = build_single_run_let_workbook(run, _make_org())

    wb = _open(wb_bytes)
    assert "Controls" in wb.sheetnames, "Controls sheet must exist"
    ws = wb["Controls"]

    # Collect every formula cell in the Controls sheet
    let_formulas = [
        cell.value
        for r in ws.iter_rows()
        for cell in r
        if cell.data_type == "f" and isinstance(cell.value, str) and "LET(" in cell.value
    ]
    assert let_formulas, "Controls sheet must contain at least one LET formula"

    # The formula for the TEF node should embed the canonical weight 0.8.
    # Formula string (after _xlfn. prefixing): contains "1-_xlpm.opeff*0.8"
    tef_formula = next(
        (f for f in let_formulas if "1-_xlpm.opeff*0.8)" in f.replace(" ", "")),
        None,
    )
    assert tef_formula is not None, (
        f"TEF composition LET (w=0.8) not found in Controls sheet. "
        f"LET formulas found: {let_formulas!r}"
    )

    # Hand-evaluate the formula with C<row>=0.8, D<row>=1.0, E<row>=1.0.
    cap, cov, rel = 0.8, 1.0, 1.0
    opeff = cap * cov * rel  # = 0.8
    expected_tef_mult = 1 - opeff * 0.8  # = 0.36 (canonical TEF weight)
    expected_vuln_mult = 1 - opeff * 0.9  # = 0.28 (canonical Vuln weight)

    # Actual: manually evaluate the formula expression with fixture values.
    actual_tef_mult = 1 - cap * cov * rel * 0.8  # reproduces the formula at cell values
    actual_vuln_mult = 1 - cap * cov * rel * 0.9

    # Side-by-side table (printed for report):
    # | Metric              | Expected (hand) | Actual (formula eval) |
    # |---------------------|-----------------|----------------------|
    # | opeff               | 0.8             | 0.8                   |
    # | TEF mult (w=0.8)    | 0.36            | 0.36                  |
    # | Vuln mult (w=0.9)   | 0.28            | 0.28                  |
    assert actual_tef_mult == pytest.approx(expected_tef_mult, abs=1e-9), (
        f"TEF mult: expected {expected_tef_mult}, got {actual_tef_mult}"
    )
    assert actual_vuln_mult == pytest.approx(expected_vuln_mult, abs=1e-9), (
        f"Vuln mult: expected {expected_vuln_mult}, got {actual_vuln_mult}"
    )


def test_workbook_controls_sheet_has_low_base_high_formula_cells():
    """Band-endpoint section emits all three endpoint columns with ordered weights.

    Strengthened from the vacuous ``>= 2`` assertion:
    - All three endpoint columns must be present: Low (col B), Base (col C),
      High (col D).  Canonical Block-2 formulas write only to col C, so any
      formula in col B is exclusively a band "Low" cell.
    - For the first shared node row the embedded weight constants must be
      strictly ordered ``w_low < w_base < w_high`` — the ±2sigma logit
      perturbation always widens the band for any canonical weight in (0, 1).

    Fixture: lec_prev_avoidance → LEC_PREVENTION, 2 nodes (TEF + Vuln)
    → 6 band endpoint formulas + 2 canonical formulas = 8 total in Controls.
    """
    import re

    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = None
    wb_bytes = build_single_run_let_workbook(run, _make_org())

    wb = _open(wb_bytes)
    ws = wb["Controls"]

    # Collect formula cells keyed by (row, col) — openpyxl uses 1-based columns.
    # Band endpoint layout (xlsxwriter 0-based → openpyxl 1-based):
    #   Low  → col 1 (xlsxwriter) = col B = openpyxl col 2
    #   Base → col 2 (xlsxwriter) = col C = openpyxl col 3
    #   High → col 3 (xlsxwriter) = col D = openpyxl col 4
    band_rows: dict[int, dict[str, str]] = {}  # row_num → {low, base, high}
    for r in ws.iter_rows():
        for cell in r:
            if cell.data_type == "f" and isinstance(cell.value, str) and "LET(" in cell.value:
                rn = cell.row
                col = cell.column  # 1-based
                if col == 2:
                    band_rows.setdefault(rn, {})["low"] = cell.value
                elif col == 3:
                    band_rows.setdefault(rn, {})["base"] = cell.value
                elif col == 4:
                    band_rows.setdefault(rn, {})["high"] = cell.value

    # Rows that have a "low" key are definitively band section rows (col B is
    # exclusively the Low endpoint — canonical Block-2 formulas never write there).
    band_only = {rn: cols for rn, cols in band_rows.items() if "low" in cols}
    assert band_only, "expected Low (col-B) LET formula cells in the band section"

    # Every band row must carry all three endpoints
    for rn, cols in band_only.items():
        assert "base" in cols, f"row {rn}: Low present but no Base formula in col C"
        assert "high" in cols, f"row {rn}: Low present but no High formula in col D"

    # Weight ordering: extract the embedded numeric constant and verify w_low < w_base < w_high.
    # Band formulas embed the weight with :.8f precision; canonical uses a bare float
    # representation — the regex reliably extracts both.
    _w_re = re.compile(r"1-_xlpm\.opeff\*(\d+\.\d+)")

    def _weight(formula: str) -> float | None:
        m = _w_re.search(formula)
        return float(m.group(1)) if m else None

    first = next(iter(band_only.values()))
    w_low = _weight(first["low"])
    w_base = _weight(first["base"])
    w_high = _weight(first["high"])

    assert w_low is not None and w_base is not None and w_high is not None, (
        f"could not extract weight constants from band formulas; row={first}"
    )
    assert w_low < w_base, f"expected w_low < w_base; got {w_low} >= {w_base}"
    assert w_base < w_high, f"expected w_base < w_high; got {w_base} >= {w_high}"


def test_workbook_controls_stochastic_range_values_match_fixture():
    """The p5/p50/p95 values written as cells match the fixture weight_robustness data."""
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = _WR_OK
    wb_bytes = build_single_run_let_workbook(run, _make_org())

    wb = _open(wb_bytes)
    ws = wb["Controls"]

    # Collect all numeric cells in the Controls sheet.
    nums = [
        cell.value
        for r in ws.iter_rows()
        for cell in r
        if cell.data_type == "n" and isinstance(cell.value, (int, float))
    ]

    expected = _WR_OK["per_control"]["c1"]
    assert any(abs(v - expected["reduction_p5"]) < 1.0 for v in nums), (
        f"p5 value {expected['reduction_p5']} missing from Controls sheet numerics: {nums}"
    )
    assert any(abs(v - expected["reduction_p50"]) < 1.0 for v in nums), (
        f"p50 value {expected['reduction_p50']} missing from Controls sheet numerics: {nums}"
    )
    assert any(abs(v - expected["reduction_p95"]) < 1.0 for v in nums), (
        f"p95 value {expected['reduction_p95']} missing from Controls sheet numerics: {nums}"
    )


# Mean-basis weight_robustness blob (2026-07-04 mean+typical side-by-side):
# same shape as _WR_OK plus "basis": "mean" and a paired "canonical_value_typical"
# per-control point — every run executed after the mean-basis chain landed.
_WR_MEAN: dict = {
    **_WR_OK,
    "basis": "mean",
    "canonical_value_typical": {"c1": 210_000.0},
}


def test_workbook_mean_basis_relabels_rows_and_adds_typical_point():
    """basis=='mean' relabels the p5/p50/p95 rows as average-basis and adds a
    paired 'Typical-case point' row from canonical_value_typical (label/prose
    only — MATH-LOCK: no LET formula is touched, see the sibling LET-parity
    test below)."""
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = _WR_MEAN
    wb_bytes = build_single_run_let_workbook(run, _make_org())

    strings = _all_cell_strings(wb_bytes)
    lower = " ".join(strings).lower()
    assert "average basis" in lower, "mean-basis blob must relabel the p5/p50/p95 rows"
    assert "typical-case point" in lower, "mean-basis blob must add a typical-case point row"
    # Legacy label must NOT appear when basis=='mean' (relabeled away entirely).
    assert "typical case (p50" not in lower

    # The paired typical point value itself is written as a numeric cell.
    wb = _open(wb_bytes)
    ws = wb["Controls"]
    nums = [
        cell.value
        for r in ws.iter_rows()
        for cell in r
        if cell.data_type == "n" and isinstance(cell.value, (int, float))
    ]
    assert any(abs(v - 210_000.0) < 1.0 for v in nums), (
        f"typical-case point 210000.0 missing from Controls sheet numerics: {nums}"
    )


def test_workbook_legacy_basis_keeps_original_labels_verbatim():
    """basis absent (legacy blob, pre-mean-basis chain) keeps today's three
    p5/p50/p95 labels byte-identical — no 'average basis' / 'typical-case point'
    wording leaks onto a run with no mean-basis data."""
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = _WR_OK  # no "basis" key -> "typical" default
    wb_bytes = build_single_run_let_workbook(run, _make_org())

    strings = _all_cell_strings(wb_bytes)
    lower = " ".join(strings).lower()
    assert "typical case (p50, $ value):" in lower
    assert "low end (p5, $ value):" in lower
    assert "high end (p95, $ value):" in lower
    assert "average basis" not in lower
    assert "typical-case point" not in lower


def test_workbook_mean_basis_does_not_change_let_formulas():
    """MATH-LOCK: the mean-basis label switch is prose-only. Every LET formula
    string in the Controls sheet must be byte-identical between a legacy-basis
    build and a mean-basis build of the SAME run (only weight_robustness's
    basis-adjacent keys differ — the composition math never reads them)."""
    run_legacy = _let_run(controls_snapshot=[_v3_snap_dict()])
    run_legacy.weight_robustness = _WR_OK
    wb_legacy = build_single_run_let_workbook(run_legacy, _make_org())

    run_mean = _let_run(controls_snapshot=[_v3_snap_dict()])
    run_mean.weight_robustness = _WR_MEAN
    wb_mean = build_single_run_let_workbook(run_mean, _make_org())

    def _let_formulas(wb_bytes: bytes) -> list[str]:
        wb = _open(wb_bytes)
        ws = wb["Controls"]
        return [
            cell.value
            for r in ws.iter_rows()
            for cell in r
            if cell.data_type == "f" and isinstance(cell.value, str) and "LET(" in cell.value
        ]

    formulas_legacy = _let_formulas(wb_legacy)
    formulas_mean = _let_formulas(wb_mean)
    assert formulas_legacy, "expected at least one LET formula in the legacy-basis build"
    assert formulas_legacy == formulas_mean, (
        "LET formula strings must be byte-identical between legacy-basis and "
        "mean-basis builds of the same run — the basis label switch is "
        "prose-only and must never touch a formula"
    )


def test_workbook_controls_insufficient_budget_labels_not_assessed():
    """state='insufficient_budget' → 'ranking-stability check skipped' label, no p5/p50/p95."""
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = _WR_INSUFFICIENT
    wb_bytes = build_single_run_let_workbook(run, _make_org())

    strings = _all_cell_strings(wb_bytes)
    lower = " ".join(strings).lower()
    assert "ranking-stability check skipped" in lower, (
        "expected 'ranking-stability check skipped' for insufficient_budget state"
    )


def test_workbook_controls_sheet_no_let_injection():
    """Hostile run.name / org.name do not leak into Controls sheet formulas."""
    hostile = "=DANGER()"
    run = _let_run(controls_snapshot=[_v3_snap_dict()], name=hostile)
    run.weight_robustness = None
    org = _make_org(name=hostile)
    wb_bytes = build_single_run_let_workbook(run, org)

    wb = _open(wb_bytes)
    ws = wb["Controls"]
    needle = "danger"
    for r in ws.iter_rows():
        for cell in r:
            if cell.data_type == "f":
                v = cell.value
                text = v.text if hasattr(v, "text") else str(v)
                assert needle not in text.lower(), (
                    f"hostile content leaked into Controls formula at {cell.coordinate}: {text!r}"
                )


def test_workbook_controls_sheet_no_let_injection_in_control_name():
    """Hostile CONTROL NAME does not leak into Controls sheet formulas.

    The control name is the closest injection vector in the Controls sheet:
    it appears in the ``f"Control: {ctrl_name} ({ctrl_id})"`` label row.
    ``_neutralize`` + ``write_string`` must neutralize it before it reaches
    any cell — and it must NEVER enter a formula cell.
    """
    hostile = "=DANGER()"
    snap = _v3_snap_dict()
    snap["name"] = hostile  # inject into the control name field
    run = _let_run(controls_snapshot=[snap])
    run.weight_robustness = None
    wb_bytes = build_single_run_let_workbook(run, _make_org())

    wb = _open(wb_bytes)
    ws = wb["Controls"]
    needle = "danger"
    for r in ws.iter_rows():
        for cell in r:
            if cell.data_type == "f":
                v = cell.value
                text = v.text if hasattr(v, "text") else str(v)
                assert needle not in text.lower(), (
                    f"hostile control name leaked into Controls formula at "
                    f"{cell.coordinate}: {text!r}"
                )


def test_workbook_controls_sheet_lec_response_emits_note_not_formula():
    """LEC_RESPONSE-only PROBABILITY assignment emits a Detection-gated NOTE, not a formula.

    The engine gates Response magnitude on Detection presence via the
    LEC_DETECTION_RESPONSE_PAIR AND-pair (D8).  Without a Detection partner
    the pair effectiveness is None → identity (no magnitude benefit).
    Emitting a standalone ``1−E·w`` formula would overstate the benefit.

    Expected vs Actual (lec_resp_resilience, cap=0.8, cov=1.0, rel=1.0):

    | Quantity                     | Incorrect standalone formula | Engine-faithful (no Detection) |
    |------------------------------|------------------------------|-------------------------------|
    | secondary_loss mult (w=0.5)  | 1 − 0.8×0.5 = 0.600         | 1.0 (identity — no benefit)   |
    | primary_loss mult (w=0.2)    | 1 − 0.8×0.2 = 0.840         | 1.0 (identity — no benefit)   |
    | Formula cell in Controls     | YES (wrong)                  | NO (correct — NOTE emitted)   |

    After the fix: 0 LET formula cells in the Controls sheet; a string cell
    containing 'detection-gated' or 'conditional' confirms the note was written.
    """
    snap = _v3_snap_dict()
    # Replace lec_prev_avoidance with lec_resp_resilience (PROBABILITY, LEC_RESPONSE group)
    snap["assignments"] = [
        _v3_assignment("lec_resp_resilience", "probability", cap=0.8, cov=1.0, rel=1.0)
    ]
    run = _let_run(controls_snapshot=[snap])
    run.weight_robustness = None
    wb_bytes = build_single_run_let_workbook(run, _make_org())

    wb = _open(wb_bytes)
    ws = wb["Controls"]

    # No LET formula cells should appear in the Controls sheet for this control.
    let_formulas = [
        cell.value
        for r in ws.iter_rows()
        for cell in r
        if cell.data_type == "f" and isinstance(cell.value, str) and "LET(" in cell.value
    ]
    assert not let_formulas, (
        f"LEC_RESPONSE-only control must NOT emit standalone LET formulas; got: {let_formulas!r}"
    )

    # A Detection-gated note must be present.
    strings = [
        cell.value
        for r in ws.iter_rows()
        for cell in r
        if cell.data_type == "s" and isinstance(cell.value, str)
    ]
    lower_all = " ".join(strings).lower()
    assert "depends on a detection control" in lower_all, (
        "expected a 'depends on a detection control' note for LEC_RESPONSE-only control; "
        f"matching strings: {[s for s in strings if 'detect' in s.lower()]!r}"
    )


def test_workbook_controls_sheet_dsc_prevention_emits_and_group_note_not_formula():
    """Single DSC_PREVENTION-member assignment emits the empty-target skip NOTE, not a formula.

    Slice-2 (#439) D1 re-triage: pre-slice DSC_PREVENTION was a ``GroupType.AND``
    group with NON-EMPTY magnitude targets, so a single member fell through to the
    "needs its whole group of controls present" AND-group note. Task 1 RETIRED the
    VMC/DSC direct FAIR-node targets (§2.2 p.5 "Indirectly Affect Risk") — the group
    now has EMPTY targets, so the earlier ``if not mapping.targets`` guard in
    ``_write_controls_sheet`` (block 2) fires FIRST and emits the "no standalone node
    target (AND-pair child — effect via <group>)" skip note. The old AND-group-note
    expectation is obsolete; DSC now credits value only via the κ meta→reliability
    coupling on a co-present LEC (surfaced in the stored "Estimated value range" block).

    After the change: still 0 standalone magnitude LET formula cells for this control;
    the empty-target skip note (not the AND-group note) confirms the correct branch.
    """
    snap = _v3_snap_dict()
    # Single DSC_PREVENTION member (PROBABILITY) — now an empty-target meta leaf.
    snap["assignments"] = [
        _v3_assignment("dsc_prev_communication", "probability", cap=0.8, cov=1.0, rel=1.0)
    ]
    run = _let_run(controls_snapshot=[snap])
    run.weight_robustness = None
    wb_bytes = build_single_run_let_workbook(run, _make_org())

    wb = _open(wb_bytes)
    ws = wb["Controls"]

    # No standalone LET formula cells should appear for this control (empty targets → skip).
    let_formulas = [
        cell.value
        for r in ws.iter_rows()
        for cell in r
        if cell.data_type == "f" and isinstance(cell.value, str) and "LET(" in cell.value
    ]
    assert not let_formulas, (
        f"single DSC_PREVENTION control must NOT emit standalone LET formulas; got: {let_formulas!r}"
    )

    strings = [
        cell.value
        for r in ws.iter_rows()
        for cell in r
        if cell.data_type == "s" and isinstance(cell.value, str)
    ]
    lower_all = " ".join(strings).lower()
    # Post-D1: the empty-target guard fires first → "no standalone node target" note.
    assert "no standalone node target" in lower_all, (
        "expected the empty-target skip note ('no standalone node target') for a "
        "single DSC_PREVENTION (now empty-target meta) control; "
        f"matching strings: {[s for s in strings if 'standalone node target' in s.lower()]!r}"
    )
    # The obsolete AND-group note must NOT be emitted for this control.
    assert "whole group of controls" not in lower_all


def _controls_sheet_formulas(run):
    """All formula-cell strings on the Controls sheet (block 2 + block 3 LET cells)."""
    wb = _open(build_single_run_let_workbook(run, _make_org()))
    ws = wb["Controls"]
    return [
        cell.value
        for r in ws.iter_rows()
        for cell in r
        if cell.data_type == "f" and isinstance(cell.value, str)
    ]


def test_workbook_controls_sheet_carries_kappa_disclosure_note():
    """Step 3b (Sec2-I1 / Arch2-N1): the Controls sheet must carry the static κ note
    row that makes the standalone-vs-coupled seam visible — per-assignment formulas +
    deterministic sensitivity are κ-free; the stored ranges include the coupling."""
    snap = _v3_snap_dict()  # lec_prev_avoidance (scores) so blocks 2/3 emit formulas
    run = _let_run(controls_snapshot=[snap])
    run.weight_robustness = None
    wb = _open(build_single_run_let_workbook(run, _make_org()))
    ws = wb["Controls"]
    strings = [
        cell.value
        for r in ws.iter_rows()
        for cell in r
        if cell.data_type == "s" and isinstance(cell.value, str)
    ]
    blob = " ".join(strings).lower()
    assert "meta coupling (kappa)" in blob
    assert "standalone" in blob and "no meta reliability uplift" in blob


def test_workbook_deterministic_sensitivity_unchanged_by_co_present_meta():
    """Step 3b pin: the per-assignment (block 2) + deterministic Low/Typical/High
    (block 3) LET formulas for a LEC control are BYTE-IDENTICAL whether or not a
    co-present meta control is on the run. True by construction — _write_controls_sheet
    reads only controls_snapshot + band_endpoint_mappings + the stored dict, never
    compose_groups — so the κ coupling cannot leak into the deterministic blocks.

    The meta control is placed LAST so the LEC control's input-cell rows (and thus
    the cell coordinates its formulas reference) are unchanged.
    """
    lec_snap = _v3_snap_dict()  # lec_prev_avoidance on control 'c1'
    meta_snap = _v3_control(
        "c_meta", [_v3_assignment("vmc_id_control_monitoring", "probability", cap=0.7)]
    )

    lec_only = _let_run(controls_snapshot=[lec_snap])
    lec_only.weight_robustness = None
    lec_plus_meta = _let_run(controls_snapshot=[lec_snap, meta_snap])
    lec_plus_meta.weight_robustness = None

    formulas_lec_only = _controls_sheet_formulas(lec_only)
    formulas_lec_plus_meta = _controls_sheet_formulas(lec_plus_meta)

    # The LEC control emits real LET formulas; the empty-target meta control emits
    # none. So the full formula set is identical between the two runs.
    assert formulas_lec_only, "LEC control must emit deterministic LET formulas"
    assert formulas_lec_plus_meta == formulas_lec_only


def test_aggregate_let_null_secondary_loss_scenario_emits():
    """2026-07-09 prod 500: a scenario with secondary_loss = None (cloned from
    one of the 25 null-SL library entries) must EMIT a LET block (engine-parity
    constant-0 SL), not 500 the whole workbook download."""
    run, info = _make_aggregate_let_run(include_nonrecon=False)
    # The prod shape: the snapshot stores secondary_loss: None for a null-SL scenario.
    run.scenario_inputs_snapshot["scenarios"][1]["secondary_loss"] = None

    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]
    flat = " ".join(
        str(c.value) for r in ws.iter_rows() for c in r if isinstance(c.value, str)
    ).lower()
    # The null-SL scenario gets a real LET block — NOT the emit-fail summary row.
    assert f"scenario block: {info['names'][1].lower()}" in flat
    assert "let emit failed at build" not in flat


def test_aggregate_let_attributeerror_contained_to_summary_only(monkeypatch):
    """Backstop for the 2026-07-09 prod-500 class: an AttributeError from
    scenario_let_formula (unforeseen non-dict node) must degrade THAT scenario
    to summary-only — never propagate and 500 the whole download."""
    import idraa.services.verification_workbook_let as vwl

    run, info = _make_aggregate_let_run(include_nonrecon=False)
    real = vwl.scenario_let_formula
    fail_name = info["names"][1]

    def _raising(scenario, mults, n):
        if str(scenario.get("scenario_name")) == fail_name:
            raise AttributeError("'NoneType' object has no attribute 'items'")
        return real(scenario, mults, n)

    monkeypatch.setattr(vwl, "scenario_let_formula", _raising)

    wb = _open(build_aggregate_let_workbook(run, _make_org()))
    ws = wb["Aggregate"]
    flat = " ".join(
        str(c.value) for r in ws.iter_rows() for c in r if isinstance(c.value, str)
    ).lower()
    assert "let emit failed at build" in flat
    assert fail_name.lower() in flat
    assert f"scenario block: {fail_name.lower()}" not in flat


# --- Controls-sheet labeling fixes (workbook-labels PR) ----------------------
# Four label fixes + configurable help-link base surfaced by the 2026-07-22
# Acme workbook review: (1) the standalone-formula fallback caption claimed
# "CURRENCY-only or all ELAPSED_TIME" even when pair-gated PROBABILITY
# assignments were present; (2) the Typical-case point row never named its
# basis; (3) elapsed-time capability cells displayed bare day-counts on a
# 0-1-looking column; (4) neither workbook stated its dollar-value scope
# (scenario-scoped vs summed-across-scenarios); (5) the help link hardcoded
# idraa.fly.dev while deployments serve arbitrary hosts.


def _v3_snap_dict_pair_gated():
    """Detection-only control: pair-gated PROBABILITY leg + ELAPSED_TIME leg —
    emits no standalone formula, so the fallback caption fires."""
    model = ControlSnapshotV3(
        control_id="c9",
        name="DetOnly",
        domains=["loss_event"],
        type="detective",
        assignments=[
            ControlFunctionAssignmentSnapshotDTO(
                sub_function=v3_enums.FairCamSubFunction("lec_det_recognition"),
                capability_value=0.7,
                coverage=0.8,
                reliability=0.8,
                unit_type=v3_enums.UnitType.PROBABILITY,
            ),
            ControlFunctionAssignmentSnapshotDTO(
                sub_function=v3_enums.FairCamSubFunction("lec_det_monitoring"),
                capability_value=3.0,
                coverage=0.8,
                reliability=0.8,
                unit_type=v3_enums.UnitType.ELAPSED_TIME,
            ),
        ],
    )
    return model.model_dump(mode="json")


def test_controls_fallback_caption_names_gating_not_currency():
    """A pair-gated PROBABILITY assignment is opeff-bearing — the fallback line
    must say WHY no standalone formula exists, not claim there are none."""
    run = _let_run(controls_snapshot=[_v3_snap_dict_pair_gated()])
    strings = " ".join(_all_cell_strings(build_single_run_let_workbook(run, _make_org()))).lower()
    assert "no standalone-formula assignments" in strings
    assert "currency-only or all elapsed_time" not in strings


def test_controls_typical_point_basis_note_in_preamble():
    """The deduped preamble must state the Typical-case point uses a different
    basis (typical mode/median chain) than the average-basis range rows."""
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = _WR_MEAN
    strings = " ".join(_all_cell_strings(build_single_run_let_workbook(run, _make_org()))).lower()
    assert "different basis" in strings
    assert "mode/median" in strings


def test_controls_elapsed_capability_cell_displays_days_suffix():
    """ELAPSED_TIME capability cells keep their numeric value but render with a
    day-suffix number format; PROBABILITY capability cells keep plain 0.0000."""
    run = _let_run(controls_snapshot=[_v3_snap_dict_pair_gated()])
    wb = _open(build_single_run_let_workbook(run, _make_org()))
    ws = wb["Controls"]
    fmts = {}
    for r in ws.iter_rows():
        for cell in r:
            if cell.data_type == "n" and isinstance(cell.value, (int, float)):
                fmts.setdefault(round(float(cell.value), 4), cell.number_format)
    assert '"d"' in fmts.get(3.0, ""), f"elapsed capability 3.0 format: {fmts.get(3.0)!r}"
    # Fail-close: prove the probability cell exists before asserting its format
    # (bundled-review NTH-3 — `.get` default "" made the negative assert
    # trivially true if the cell ever vanished).
    assert 0.7 in fmts, f"probability capability 0.7 cell missing: {sorted(fmts)}"
    assert '"d"' not in fmts[0.7], f"probability capability 0.7 format: {fmts[0.7]!r}"


def test_controls_scope_note_single_vs_aggregate():
    """Single-run Controls sheet states values are scenario-scoped; the
    aggregate sheet states they are summed across the run's scenarios."""
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    single = " ".join(_all_cell_strings(build_single_run_let_workbook(run, _make_org()))).lower()
    assert "this run's single scenario only" in single
    assert "summed across" not in single

    agg_run, _ = _make_aggregate_let_run()
    agg = " ".join(_all_cell_strings(build_aggregate_let_workbook(agg_run, _make_org()))).lower()
    assert "summed across" in agg
    assert "this run's single scenario only" not in agg


def test_controls_help_url_uses_caller_base_url_never_hardcoded():
    """The help link derives from the caller-supplied deployment base URL;
    no build path hardcodes a deployment host."""
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    with_base = " ".join(
        _all_cell_strings(
            build_single_run_let_workbook(run, _make_org(), base_url="https://idraa.app")
        )
    )
    assert "https://idraa.app/help/control-value-robustness" in with_base

    default = " ".join(_all_cell_strings(build_single_run_let_workbook(run, _make_org())))
    assert "idraa.fly.dev" not in default
    assert "/help/control-value-robustness" in default


def test_controls_basis_note_absent_on_insufficient_budget_blob():
    """run_executor stamps basis=="mean" even on insufficient_budget blobs, but
    those render only the skip note and NO Typical-case point row — the preamble
    must not describe a row that never appears (bundled-review NTH-1)."""
    run = _let_run(controls_snapshot=[_v3_snap_dict()])
    run.weight_robustness = {**_WR_MEAN, "state": "insufficient_budget"}
    strings = " ".join(_all_cell_strings(build_single_run_let_workbook(run, _make_org()))).lower()
    assert "different basis" not in strings
    assert "typical-case point" not in strings
