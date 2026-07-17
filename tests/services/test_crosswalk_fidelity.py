"""Re-parse the source .xlsx INDEPENDENTLY of the converter and assert the
committed JSON faithfully transcribes the X-marks (gate M6).

Tests may use openpyxl (it's dev-only; forbidden only in runtime/migrations —
see tests/arch/test_excel_writer_confined.py). We deliberately do NOT import
``scripts.build_crosswalk_seed`` — the whole point is an INDEPENDENT re-parse so
a converter bug can't make this test agree with the converter's own output.

The count-parity tests are the primary fidelity guarantee (independent X-mark
count == committed JSON link count, per framework). The hand-verified
spot-checks catch column-binding errors the bulk counts can't, and the
trailing-zero regression guards the CIS float-code rendering logic.
"""

import json
from pathlib import Path

import openpyxl
import pytest

import idraa

ROOT = Path(idraa.__file__).resolve().parent.parent.parent

# Both source xlsx files are GITIGNORED (licensing) — present on the original
# curation machine, absent in fresh clones/worktrees/CI. CIS has been gitignored
# since #350; the NIST xlsx was removed from tracking in #557 (FAIR Institute
# CC-BY-NC-ND). The parity re-parse needs the real file, so its absence is a SKIP
# (environmental), not a failure. The committed-JSON count pins live in their own
# always-running tests (test_{cis,nist}_json_link_count_pinned) so JSON
# regressions are caught wherever the suite runs.
_CIS_XLSX = ROOT / "docs" / "CIS 8.0 to FAIR-CAM Mapping V1.0.xlsx"
_NIST_XLSX = ROOT / "docs" / "reference" / "NIST CSF 1.1 to FAIR-CAM 1.0 Mapping_Final.xlsx"


def _json():
    return json.loads((ROOT / "data" / "seed_framework_crosswalk.json").read_text())["entries"]


def _cis_base_link_count():
    """CIS FAIR-Institute base link count — a plain sum (#449).

    Idraa crosswalk-seed extensions (methodology decisions; #437 rollout T1
    CIS 7.3/7.4 -> resistance, T2 CIS 4.8 -> avoidance and 14.2/16.1 ->
    resistance) live in the structurally separate ``riskflow_extension_functions``
    overlay array, so ``fair_cam_functions`` is byte-identical to the source
    X-marks and needs no subtraction arithmetic.
    """
    return sum(len(e["fair_cam_functions"]) for e in _json() if e["framework"] == "cis")


def _count_xmarks_nist(path, sheet, data_start, code_col, func_cols):
    """Count X-marks in the mapped function columns of real data rows.

    NIST codes are STRINGS ('ID.AM-1: ...'), so a data row is one whose code
    cell is a non-empty string — this mirrors the converter's row-inclusion for
    NIST (``_code_str`` returns the stripped string).
    """
    wb = openpyxl.load_workbook(ROOT / path, read_only=True, data_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    n = 0
    for r in rows[data_start:]:
        cc = r[code_col] if code_col < len(r) else None
        if not (isinstance(cc, str) and cc.strip()):
            continue
        n += sum(
            1
            for c in func_cols
            if c < len(r) and isinstance(r[c], str) and r[c].strip().upper() == "X"
        )
    return n


def _count_xmarks_cis(path, sheet, data_start, code_col, func_cols):
    """Count X-marks in the mapped function columns of real CIS data rows.

    CIS sub-category code cells are NUMERIC (floats: 1.1, 3.11, ...), not
    strings — so the NIST ``isinstance(cc, str)`` predicate would exclude EVERY
    CIS data row and yield 0. We mirror the converter's actual row-inclusion
    predicate (``scripts.build_crosswalk_seed._code_str`` returns a value for a
    non-empty float OR a non-empty string): a CIS data row is one whose code
    cell is a non-empty numeric (int/float, excluding bool) or a non-empty
    string. This keeps the counter INDEPENDENT of the converter module while
    counting the same rows it counts.
    """
    wb = openpyxl.load_workbook(ROOT / path, read_only=True, data_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    n = 0
    for r in rows[data_start:]:
        cc = r[code_col] if code_col < len(r) else None
        is_code = (isinstance(cc, (int, float)) and not isinstance(cc, bool)) or (
            isinstance(cc, str) and cc.strip()
        )
        if not is_code:
            continue
        n += sum(
            1
            for c in func_cols
            if c < len(r) and isinstance(r[c], str) and r[c].strip().upper() == "X"
        )
    return n


def test_nist_link_count_parity():
    if not _NIST_XLSX.exists():
        pytest.skip(
            "gitignored NIST source xlsx absent (#557 licensing removal) — "
            "re-parse needs the local file"
        )
    nist_cols = list(range(7, 16)) + list(range(17, 23)) + list(range(24, 34))  # 25 cols
    assert len(nist_cols) == 25
    expected = _count_xmarks_nist(
        "docs/reference/NIST CSF 1.1 to FAIR-CAM 1.0 Mapping_Final.xlsx",
        "NIST CSF 1.1 to FAIR-CAM",
        5,
        2,
        nist_cols,
    )
    got = sum(len(e["fair_cam_functions"]) for e in _json() if e["framework"] == "nist_csf")
    assert got == expected and got > 0
    assert got == 190  # pinned: matches committed JSON + independent re-parse


def test_nist_json_link_count_pinned():
    """#557: committed-JSON side of the NIST pin, split out so it runs even where
    the now-gitignored source xlsx is absent (fresh clones / worktrees / CI) —
    mirrors test_cis_json_link_count_pinned."""
    got = sum(len(e["fair_cam_functions"]) for e in _json() if e["framework"] == "nist_csf")
    assert got == 190


def test_cis_link_count_parity():
    if not _CIS_XLSX.exists():
        pytest.skip(
            "gitignored CIS source xlsx absent (issue #350) — re-parse needs the local file"
        )
    cis_cols = list(range(12, 21)) + list(range(22, 28)) + list(range(29, 39))  # 25 cols
    assert len(cis_cols) == 25
    expected = _count_xmarks_cis(
        "docs/CIS 8.0 to FAIR-CAM Mapping V1.0.xlsx",
        "CIS v8.0 to FAIR-CAM",
        9,
        2,
        cis_cols,
    )
    # Source-parity compares the xlsx X-marks to the FAIR-Institute BASE links only;
    # Idraa crosswalk-seed extensions are additive and excluded (see helper).
    got = _cis_base_link_count()
    assert got == expected and got > 0
    assert got == 283  # pinned: FAIR-Institute source X-marks (extensions excluded)


def test_cis_json_link_count_pinned():
    """#350: the committed-JSON side of the CIS pin, split out so it runs even
    where the gitignored source xlsx is absent (fresh clones / worktrees).
    Extension-aware (see _cis_base_link_count): pins the FAIR-Institute base, so
    Idraa crosswalk-seed extensions (T1/T2) do not perturb the source-parity pin."""
    assert _cis_base_link_count() == 283


def test_hand_verified_spotchecks():
    """Each (code, function) pair below was CONFIRMED by opening the source
    sheet and reading the actual X-mark in the mapped column. They guard against
    a column-binding error (right counts, wrong column→function assignment) that
    the bulk count-parity tests cannot detect."""
    rows = {(e["framework"], e["code"]): e["fair_cam_functions"] for e in _json()}
    # NIST "ID.AM-1" — single X at the DSC Situational-Awareness > Data > Asset leaf (col 26):
    assert "dsc_prev_sa_data_asset" in rows[("nist_csf", "ID.AM-1")]
    # NIST "PR.AC-7" — single X at LEC > Prevention > Resistance (col 9):
    assert "lec_prev_resistance" in rows[("nist_csf", "PR.AC-7")]
    # CIS "1.1" — X at VMC Controls Monitoring (col 25) + VMC Implementation (col 27)
    # + DSC SA Data Asset (col 31):
    assert "vmc_id_control_monitoring" in rows[("cis", "1.1")]
    assert "vmc_corr_implementation" in rows[("cis", "1.1")]
    assert "dsc_prev_sa_data_asset" in rows[("cis", "1.1")]


def test_trailing_zero_cis_codes_distinct():
    """Trouble-watch regression (Task 3): 5 CIS codes — 3.10, 4.10, 8.10, 13.10,
    16.10 — are IEEE-float-equal to their ``.1`` sibling and survive in the JSON
    ONLY because the converter honors each cell's 2-decimal number_format. A
    future refactor of ``_code_str`` could silently collapse '3.10' → '3.1',
    colliding with the real 3.1 sub-category and dropping a control. Assert all
    10 codes (5 trailing-zero + 5 ``.1`` siblings) exist as DISTINCT CIS entries
    so a collapse/collision fails loudly here."""
    codes = {e["code"] for e in _json() if e["framework"] == "cis"}
    expected = {
        "3.10",
        "3.1",
        "4.10",
        "4.1",
        "8.10",
        "8.1",
        "13.10",
        "13.1",
        "16.10",
        "16.1",
    }
    missing = expected - codes
    assert not missing, f"trailing-zero CIS codes collapsed/missing: {sorted(missing)}"
    assert len(expected & codes) == 10


def test_riskflow_extension_overlay_pinned():
    """#449: the Idraa extension overlay is pinned exactly — entries, functions,
    disjointness from the base layer, and paired narrative rationale. A new
    extension (or an accidental co-mingle back into fair_cam_functions) fails
    loudly here and forces a deliberate re-pin + provenance-migration check."""
    expected = {
        ("cis", "4.8"): ["lec_prev_avoidance"],
        ("cis", "7.3"): ["lec_prev_resistance"],
        ("cis", "7.4"): ["lec_prev_resistance"],
        ("cis", "14.2"): ["lec_prev_resistance"],
        ("cis", "16.1"): ["lec_prev_resistance"],
    }
    got = {}
    for e in _json():
        ext = e.get("riskflow_extension_functions")
        narrative = e["citation"].get("riskflow_extension")
        # Overlay array and narrative rationale must appear together.
        assert bool(ext) == bool(narrative), (e["framework"], e["code"])
        if ext:
            assert not set(ext) & set(e["fair_cam_functions"]), (e["framework"], e["code"])
            got[(e["framework"], e["code"])] = ext
    assert got == expected
